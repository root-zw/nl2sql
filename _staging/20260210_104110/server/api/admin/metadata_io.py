"""
元数据批量导入导出API
支持Excel格式的模板导出和数据导入
"""

import json
import io
from datetime import datetime
from urllib.parse import quote
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from typing import List, Optional, Dict, Any
from uuid import UUID, uuid4
import asyncpg
import structlog
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

from server.models.database import (
    FieldType,
    JoinType,
    RelationshipType,
    RuleType
)

from server.middleware.auth import require_data_admin
from server.models.admin import User as AdminUser
# 字段类型自动识别
from server.utils.field_analyzer import FieldAnalyzer

logger = structlog.get_logger()
router = APIRouter()


# ============================================================================
# 常量定义（中文选项，用于Excel下拉框）
# ============================================================================

# 字段类型选项（中文）
FIELD_TYPE_OPTIONS_CN = ["维度", "度量", "时间", "标识", "空间"]
# 字段类型映射：中文 -> 英文
FIELD_TYPE_CN_TO_EN = {
    "维度": "dimension",
    "度量": "measure", 
    "时间": "timestamp",
    "标识": "identifier",
    "空间": "spatial"
}
# 字段类型映射：英文 -> 中文
FIELD_TYPE_EN_TO_CN = {v: k for k, v in FIELD_TYPE_CN_TO_EN.items()}

# 聚合方式选项（中文）
AGGREGATION_OPTIONS_CN = ["求和", "平均值", "计数", "最大值", "最小值", "去重计数"]
# 聚合方式映射：中文 -> 英文
AGGREGATION_CN_TO_EN = {
    "求和": "SUM",
    "平均值": "AVG",
    "计数": "COUNT",
    "最大值": "MAX",
    "最小值": "MIN",
    "去重计数": "COUNT_DISTINCT"
}
AGGREGATION_EN_TO_CN = {v: k for k, v in AGGREGATION_CN_TO_EN.items()}

# JOIN类型选项（中文）
JOIN_TYPE_OPTIONS_CN = ["内连接", "左连接", "右连接", "全连接"]
JOIN_TYPE_CN_TO_EN = {
    "内连接": "INNER",
    "左连接": "LEFT",
    "右连接": "RIGHT",
    "全连接": "FULL"
}
JOIN_TYPE_EN_TO_CN = {v: k for k, v in JOIN_TYPE_CN_TO_EN.items()}

# 关系类型选项（中文）
RELATIONSHIP_TYPE_OPTIONS_CN = ["一对一", "一对多", "多对多"]
RELATIONSHIP_TYPE_CN_TO_EN = {
    "一对一": "one_to_one",
    "一对多": "one_to_many",
    "多对多": "many_to_many"
}
RELATIONSHIP_TYPE_EN_TO_CN = {v: k for k, v in RELATIONSHIP_TYPE_CN_TO_EN.items()}

# 规则类型选项（中文）
RULE_TYPE_OPTIONS_CN = ["派生指标", "默认过滤", "自定义规则"]
RULE_TYPE_CN_TO_EN = {
    "派生指标": "derived_metric",
    "默认过滤": "default_filter",
    "自定义规则": "custom_instruction"
}
RULE_TYPE_EN_TO_CN = {v: k for k, v in RULE_TYPE_CN_TO_EN.items()}

# 布尔选项（中文）
BOOL_OPTIONS_CN = ["是", "否"]

# Excel样式定义
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
READONLY_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
# 启用状态的颜色（Excel标准绿色背景 - 更饱和）
ENABLED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
# 禁用状态的颜色（Excel标准红色背景 - 更饱和）
DISABLED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


async def get_db_pool():
    """获取元数据库连接"""
    from server.utils.db_pool import get_metadata_pool
    pool = await get_metadata_pool()
    async with pool.acquire() as conn:
        yield conn


# ============================================================================
# 辅助函数
# ============================================================================

def set_column_width(ws, col_idx: int, width: int):
    """设置列宽"""
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def add_dropdown(ws, col_letter: str, start_row: int, end_row: int, options: List[str], allow_blank: bool = True):
    """为列添加下拉框"""
    if not options:
        return
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=allow_blank
    )
    dv.error = "请从下拉列表中选择"
    dv.errorTitle = "无效输入"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def style_header_row(ws, num_cols: int):
    """设置表头样式"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def style_readonly_columns(ws, readonly_cols: List[int], start_row: int, end_row: int):
    """设置只读列样式（灰色背景）"""
    for row in range(start_row, end_row + 1):
        for col in readonly_cols:
            cell = ws.cell(row=row, column=col)
            cell.fill = READONLY_FILL


def apply_table_borders(ws, num_cols: int, end_row: int):
    """为表格数据区域设置边框"""
    for row in range(2, end_row + 1):  # 从第2行开始（跳过表头）
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER


def style_row_by_status(ws, row_idx: int, is_enabled: bool, num_cols: int, readonly_cols: List[int] = None):
    """根据启用状态设置整行颜色"""
    fill = ENABLED_FILL if is_enabled else DISABLED_FILL
    readonly_cols = readonly_cols or []
    
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        # 只读列使用灰色，其他列根据启用状态设置颜色
        if col in readonly_cols:
            cell.fill = READONLY_FILL
        else:
            cell.fill = fill


def add_conditional_formatting(ws, status_col: str, num_cols: int, max_row: int, readonly_cols: List[int] = None):
    """
    添加条件格式规则，根据"是否启用"列的值自动变换行颜色
    
    Args:
        ws: 工作表
        status_col: 启用状态列的字母（如 'I'）
        num_cols: 总列数
        max_row: 最大行数（预留空行）
        readonly_cols: 只读列索引列表
    """
    readonly_cols = readonly_cols or []
    
    # 获取最后一列的字母
    last_col_letter = get_column_letter(num_cols)
    
    # Excel条件格式规则：先添加的优先级更高
    # 所以只读列的规则要先添加，才能覆盖后面的启用/禁用颜色
    
    # 规则1（最高优先级）：只读列始终保持灰色
    if readonly_cols:
        for col_idx in readonly_cols:
            col_letter = get_column_letter(col_idx)
            col_range = f"{col_letter}2:{col_letter}{max_row}"
            readonly_rule = FormulaRule(
                formula=['TRUE'],  # 始终为真
                fill=READONLY_FILL,
                stopIfTrue=True  # 匹配后停止评估后续规则
            )
            ws.conditional_formatting.add(col_range, readonly_rule)
    
    # 条件格式应用范围（从第2行开始，跳过表头）
    range_str = f"A2:{last_col_letter}{max_row}"
    
    # 规则2：当状态列为"是"时，显示绿色背景
    enabled_rule = FormulaRule(
        formula=[f'${status_col}2="是"'],
        fill=ENABLED_FILL
    )
    
    # 规则3：当状态列为"否"时，显示红色背景  
    disabled_rule = FormulaRule(
        formula=[f'${status_col}2="否"'],
        fill=DISABLED_FILL
    )
    
    ws.conditional_formatting.add(range_str, enabled_rule)
    ws.conditional_formatting.add(range_str, disabled_rule)


# ============================================================================
# 导出API
# ============================================================================

@router.get("/metadata/export")
async def export_metadata_unified(
    connection_ids: Optional[str] = Query(None, description="数据源ID，支持逗号分隔多个；为空时导出所有"),
    table_name: Optional[str] = Query(None, description="指定表名，只导出该表的配置（支持逗号分隔多个表）"),
    include_domains: bool = Query(True, description="包含业务域配置"),
    include_tables: bool = Query(True, description="包含表配置"),
    include_fields: bool = Query(True, description="包含字段配置"),
    include_enums: bool = Query(True, description="包含枚举值配置"),
    include_relationships: bool = Query(True, description="包含表关系配置"),
    include_rules: bool = Query(True, description="包含全局规则配置"),
    current_user: AdminUser = Depends(require_data_admin),
    db = Depends(get_db_pool)
):
    """
    统一的元数据导出端点（Excel格式）
    
    - connection_ids: 为空导出所有数据源；单个ID或逗号分隔多个ID
    - 支持按表名筛选
    """
    try:
        # 解析数据源ID列表
        connection_id_list = []
        if connection_ids:
            connection_id_list = [UUID(cid.strip()) for cid in connection_ids.split(',') if cid.strip()]
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # 解析表名参数
        table_names_list = None
        if table_name:
            table_names_list = [t.strip() for t in table_name.split(',') if t.strip()]
        
        # 根据是否指定数据源决定导出范围
        if not connection_id_list:
            # 导出所有数据源
            domain_rows = await db.fetch(
                "SELECT domain_id, domain_code, domain_name FROM business_domains ORDER BY connection_id NULLS FIRST, sort_order"
            )
            domain_codes = [row['domain_code'] for row in domain_rows]
            domain_map = {str(row['domain_id']): row['domain_code'] for row in domain_rows}
            
            _create_instructions_sheet(wb)
            if include_domains:
                await _create_domains_sheet_all(wb, db, domain_rows)
            if include_tables:
                await _create_tables_sheet_all(wb, db, domain_codes, domain_map, table_names_list)
            if include_fields:
                await _create_fields_sheet_all(wb, db, table_names_list)
            if include_enums:
                await _create_enums_sheet_all(wb, db, table_names_list)
            if include_relationships:
                await _create_relationships_sheet_all(wb, db, table_names_list)
            if include_rules:
                await _create_rules_sheet_all(wb, db, domain_codes)
            
            filename_suffix = "all"
        elif len(connection_id_list) == 1:
            # 单个数据源
            connection_id = connection_id_list[0]
            conn_info = await db.fetchrow(
                "SELECT connection_name FROM database_connections WHERE connection_id = $1",
                connection_id
            )
            if not conn_info:
                raise HTTPException(status_code=404, detail=f"数据库连接 {connection_id} 不存在")
            
            domain_rows = await db.fetch(
                "SELECT domain_id, domain_code, domain_name FROM business_domains WHERE connection_id = $1 OR connection_id IS NULL ORDER BY sort_order",
                connection_id
            )
            domain_codes = [row['domain_code'] for row in domain_rows]
            domain_map = {str(row['domain_id']): row['domain_code'] for row in domain_rows}
            
            _create_instructions_sheet(wb)
            if include_domains:
                await _create_domains_sheet(wb, db, connection_id, domain_rows)
            if include_tables:
                await _create_tables_sheet(wb, db, connection_id, domain_codes, domain_map, table_names_list)
            if include_fields:
                await _create_fields_sheet(wb, db, connection_id, table_names_list)
            if include_enums:
                await _create_enums_sheet(wb, db, connection_id, table_names_list)
            if include_relationships:
                await _create_relationships_sheet(wb, db, connection_id, table_names_list)
            if include_rules:
                await _create_rules_sheet(wb, db, connection_id, domain_codes)
            
            filename_suffix = conn_info['connection_name'].replace(' ', '_')
        else:
            # 多个数据源 - 使用 _all 函数并传入 connection_ids 参数
            domain_rows = await db.fetch(
                "SELECT domain_id, domain_code, domain_name FROM business_domains WHERE connection_id = ANY($1) OR connection_id IS NULL ORDER BY sort_order",
                connection_id_list
            )
            domain_codes = [row['domain_code'] for row in domain_rows]
            domain_map = {str(row['domain_id']): row['domain_code'] for row in domain_rows}
            
            _create_instructions_sheet(wb)
            if include_domains:
                await _create_domains_sheet_all(wb, db, domain_rows, connection_id_list)
            if include_tables:
                await _create_tables_sheet_all(wb, db, domain_codes, domain_map, table_names_list, connection_id_list)
            if include_fields:
                await _create_fields_sheet_all(wb, db, table_names_list, connection_id_list)
            if include_enums:
                await _create_enums_sheet_all(wb, db, table_names_list, connection_id_list)
            if include_relationships:
                await _create_relationships_sheet_all(wb, db, table_names_list, connection_id_list)
            if include_rules:
                await _create_rules_sheet_all(wb, db, domain_codes, connection_id_list)
            
            filename_suffix = f"multi_{len(connection_id_list)}"
        
        # 保存到BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"metadata_{filename_suffix}_{timestamp}.xlsx"
        
        logger.info(f"导出元数据配置: {filename}")
        
        # 对文件名进行 URL 编码，支持中文文件名
        encoded_filename = quote(filename)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("导出元数据配置失败")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.get("/metadata/export/{connection_id}")
async def export_metadata_template(
    connection_id: UUID,
    table_name: Optional[str] = Query(None, description="指定表名，只导出该表的配置（支持逗号分隔多个表）"),
    include_domains: bool = Query(True, description="包含业务域配置"),
    include_tables: bool = Query(True, description="包含表配置"),
    include_fields: bool = Query(True, description="包含字段配置"),
    include_enums: bool = Query(True, description="包含枚举值配置"),
    include_relationships: bool = Query(True, description="包含表关系配置"),
    include_rules: bool = Query(True, description="包含全局规则配置"),
    current_user: AdminUser = Depends(require_data_admin),
    db = Depends(get_db_pool)
):
    """
    导出指定数据源的元数据配置模板（Excel格式）
    
    - 包含Schema同步后的表/列基础信息（只读列）
    - 包含已配置的语义信息（可编辑列）
    - 下拉框限制输入选项
    - 支持按表名筛选（table_name参数，多个表用逗号分隔）
    """
    try:
        # 获取连接信息
        conn_info = await db.fetchrow(
            "SELECT connection_name FROM database_connections WHERE connection_id = $1",
            connection_id
        )
        if not conn_info:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"数据库连接 {connection_id} 不存在"
            )
        
        wb = Workbook()
        # 删除默认sheet
        wb.remove(wb.active)
        
        # 解析表名参数（支持逗号分隔多个表）
        table_names_list = None
        if table_name:
            table_names_list = [t.strip() for t in table_name.split(',') if t.strip()]
        
        # 获取业务域列表（用于下拉框）
        domain_rows = await db.fetch(
            "SELECT domain_id, domain_code, domain_name FROM business_domains WHERE connection_id = $1 OR connection_id IS NULL ORDER BY sort_order",
            connection_id
        )
        domain_codes = [row['domain_code'] for row in domain_rows]
        domain_map = {str(row['domain_id']): row['domain_code'] for row in domain_rows}
        
        # 1. 使用说明Sheet（放在第一个）
        _create_instructions_sheet(wb)
        
        # 2. 业务域配置Sheet（不受表名筛选影响）
        if include_domains:
            await _create_domains_sheet(wb, db, connection_id, domain_rows)
        
        # 3. 表配置Sheet
        if include_tables:
            await _create_tables_sheet(wb, db, connection_id, domain_codes, domain_map, table_names_list)
        
        # 4. 字段配置Sheet
        if include_fields:
            await _create_fields_sheet(wb, db, connection_id, table_names_list)
        
        # 5. 枚举值配置Sheet
        if include_enums:
            await _create_enums_sheet(wb, db, connection_id, table_names_list)
        
        # 6. 表关系配置Sheet
        if include_relationships:
            await _create_relationships_sheet(wb, db, connection_id, table_names_list)
        
        # 7. 全局规则配置Sheet（不受表名筛选影响）
        if include_rules:
            await _create_rules_sheet(wb, db, connection_id, domain_codes)
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名（对中文进行URL编码）
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"metadata_{conn_info['connection_name']}_{timestamp}.xlsx"
        # URL编码文件名以支持中文
        encoded_filename = quote(filename, safe='')
        
        logger.info(f"导出元数据模板成功", connection_id=str(connection_id), filename=filename)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("导出元数据模板失败")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"导出失败: {str(e)}"
        )


async def _create_domains_sheet(wb: Workbook, db, connection_id: UUID, domain_rows: list):
    """创建业务域配置Sheet"""
    ws = wb.create_sheet("业务域配置")
    
    # 表头
    headers = ["域代码", "域名称", "描述", "关键词(逗号分隔)", "典型查询(逗号分隔)", "图标", "颜色", "排序", "是否启用"]
    ws.append(headers)
    style_header_row(ws, len(headers))
    
    # 填充数据
    row_idx = 2
    for row in domain_rows:
        # 获取完整信息
        full_domain = await db.fetchrow(
            "SELECT * FROM business_domains WHERE domain_id = $1",
            row['domain_id']
        )
        is_active = full_domain['is_active']
        ws.append([
            full_domain['domain_code'],
            full_domain['domain_name'],
            full_domain['description'] or '',
            ','.join(full_domain['keywords'] or []),
            ','.join(full_domain['typical_queries'] or []),
            full_domain['icon'] or '',
            full_domain['color'] or '#409eff',
            full_domain['sort_order'] or 0,
            '是' if is_active else '否'
        ])
        # 根据启用状态设置行颜色
        style_row_by_status(ws, row_idx, is_active, len(headers))
        row_idx += 1
    
    # 设置列宽
    col_widths = [15, 20, 30, 30, 40, 15, 12, 8, 12]
    for i, w in enumerate(col_widths, 1):
        set_column_width(ws, i, w)
    
    # 添加下拉框（使用中文选项）
    data_rows = len(domain_rows) + 100  # 预留空行
    add_dropdown(ws, 'I', 2, data_rows, BOOL_OPTIONS_CN, allow_blank=False)
    
    # 添加条件格式：根据"是否启用"列自动变换颜色
    add_conditional_formatting(ws, 'I', len(headers), data_rows)
    
    # 设置表格边框
    apply_table_borders(ws, len(headers), row_idx - 1)


async def _create_tables_sheet(wb: Workbook, db, connection_id: UUID, domain_codes: list, domain_map: dict, table_names: Optional[List[str]] = None):
    """创建表配置Sheet"""
    ws = wb.create_sheet("表配置")
    
    # 表头：只读列 + 可编辑列
    headers = ["表名(只读)", "Schema(只读)", "显示名称", "描述", "所属业务域", "标签(逗号分隔)", "数据年份", "是否启用"]
    ws.append(headers)
    style_header_row(ws, len(headers))
    
    # 构建查询条件
    query = """
        SELECT table_id, schema_name, table_name, display_name, description,
               domain_id, tags, data_year, is_included
        FROM db_tables
        WHERE connection_id = $1
    """
    params = [connection_id]
    
    if table_names:
        query += " AND table_name = ANY($2::text[])"
        params.append(table_names)
    
    query += " ORDER BY schema_name, table_name"
    
    tables = await db.fetch(query, *params)
    
    # 填充数据（使用中文布尔值）
    row_idx = 2
    for row in tables:
        domain_code = domain_map.get(str(row['domain_id']), '') if row['domain_id'] else ''
        is_enabled = row['is_included']
        ws.append([
            row['table_name'],
            row['schema_name'] or 'dbo',
            row['display_name'] or '',
            row['description'] or '',
            domain_code,
            ','.join(row['tags'] or []),
            row['data_year'] or '',
            '是' if is_enabled else '否'
        ])
        # 根据启用状态设置行颜色（只读列1,2保持灰色）
        style_row_by_status(ws, row_idx, is_enabled, len(headers), readonly_cols=[1, 2])
        row_idx += 1
    
    # 设置列宽
    col_widths = [25, 15, 20, 35, 15, 25, 12, 12]
    for i, w in enumerate(col_widths, 1):
        set_column_width(ws, i, w)
    
    # 添加下拉框（使用中文选项）
    data_rows = len(tables) + 100
    if domain_codes:
        add_dropdown(ws, 'E', 2, data_rows, domain_codes)
    add_dropdown(ws, 'H', 2, data_rows, BOOL_OPTIONS_CN, allow_blank=False)
    
    # 添加条件格式：根据"是否启用"列自动变换颜色（只读列1,2保持灰色）
    add_conditional_formatting(ws, 'H', len(headers), data_rows, readonly_cols=[1, 2])
    
    # 设置表格边框
    apply_table_borders(ws, len(headers), row_idx - 1)


async def _create_fields_sheet(wb: Workbook, db, connection_id: UUID, table_names: Optional[List[str]] = None):
    """创建字段配置Sheet"""
    ws = wb.create_sheet("字段配置")
    
    # 表头（移除了维度类型列）
    headers = [
        "表名(只读)", "Schema(只读)", "列名(只读)", "数据类型(只读)",
        "显示名称", "字段类型", "同义词(逗号分隔)",
        "默认聚合", "单位", "是否启用", "显示在明细", "描述"
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))
    
    # 构建查询条件
    query = """
        SELECT 
            t.table_name, t.schema_name, c.column_name, c.data_type,
            f.display_name, f.field_type, f.synonyms,
            f.default_aggregation, f.unit, f.is_active, f.show_in_detail, f.description
        FROM db_tables t
        JOIN db_columns c ON t.table_id = c.table_id
        LEFT JOIN fields f ON c.column_id = f.source_column_id
        WHERE t.connection_id = $1 AND t.is_included = TRUE
    """
    params = [connection_id]
    
    # 按表名筛选
    if table_names:
        query += " AND t.table_name = ANY($2::text[])"
        params.append(table_names)
    
    query += " ORDER BY t.schema_name, t.table_name, c.ordinal_position NULLS LAST"
    
    fields = await db.fetch(query, *params)
    
    # 填充数据（使用中文显示）
    row_idx = 2
    for row in fields:
        # 字段类型转中文
        field_type_cn = FIELD_TYPE_EN_TO_CN.get(row['field_type'], '') if row['field_type'] else ''
        # 聚合方式转中文
        agg_cn = AGGREGATION_EN_TO_CN.get(row['default_aggregation'], '') if row['default_aggregation'] else ''
        # 启用状态
        is_enabled = row['is_active'] is None or row['is_active']
        
        ws.append([
            row['table_name'],
            row['schema_name'] or 'dbo',
            row['column_name'],
            row['data_type'],
            row['display_name'] or '',
            field_type_cn,
            ','.join(row['synonyms'] or []) if row['synonyms'] else '',
            agg_cn,
            row['unit'] or '',
            '是' if is_enabled else '否',
            '是' if row['show_in_detail'] is None or row['show_in_detail'] else '否',
            row['description'] or ''
        ])
        # 根据启用状态设置行颜色（只读列1,2,3,4保持灰色）
        style_row_by_status(ws, row_idx, is_enabled, len(headers), readonly_cols=[1, 2, 3, 4])
        row_idx += 1
    
    # 设置列宽
    col_widths = [25, 12, 25, 15, 18, 12, 30, 14, 10, 12, 12, 35]
    for i, w in enumerate(col_widths, 1):
        set_column_width(ws, i, w)
    
    # 添加下拉框（使用中文选项）
    data_rows = len(fields) + 100
    add_dropdown(ws, 'F', 2, data_rows, FIELD_TYPE_OPTIONS_CN)  # 字段类型
    add_dropdown(ws, 'H', 2, data_rows, AGGREGATION_OPTIONS_CN)  # 默认聚合
    add_dropdown(ws, 'J', 2, data_rows, BOOL_OPTIONS_CN, allow_blank=False)  # 是否启用
    add_dropdown(ws, 'K', 2, data_rows, BOOL_OPTIONS_CN, allow_blank=False)  # 显示在明细
    
    # 添加条件格式：根据"是否启用"列自动变换颜色（只读列1,2,3,4保持灰色）
    add_conditional_formatting(ws, 'J', len(headers), data_rows, readonly_cols=[1, 2, 3, 4])
    
    # 设置表格边框
    apply_table_borders(ws, len(headers), row_idx - 1)


async def _create_enums_sheet(wb: Workbook, db, connection_id: UUID, table_names: Optional[List[str]] = None):
    """创建枚举值配置Sheet"""
    ws = wb.create_sheet("枚举值配置")
    
    # 表头
    headers = [
        "表名(只读)", "列名(只读)", "原始值", "显示值", 
        "同义词(逗号分隔)", "包含值(逗号分隔)", "是否启用"
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))
    
    # 构建查询条件
    query = """
        SELECT 
            t.table_name, c.column_name,
            e.original_value, e.display_value, e.synonyms, e.includes_values, e.is_active
        FROM field_enum_values e
        JOIN fields f ON e.field_id = f.field_id
        JOIN db_columns c ON f.source_column_id = c.column_id
        JOIN db_tables t ON c.table_id = t.table_id
        WHERE t.connection_id = $1
    """
    params = [connection_id]
    
    if table_names:
        query += " AND t.table_name = ANY($2::text[])"
        params.append(table_names)
    
    query += " ORDER BY t.table_name, c.column_name, e.frequency DESC"
    
    enums = await db.fetch(query, *params)
    
    # 填充数据
    row_idx = 2
    for row in enums:
        includes_values = row['includes_values']
        if includes_values and isinstance(includes_values, list):
            includes_str = ','.join(includes_values)
        else:
            includes_str = ''
        
        is_enabled = row['is_active']
        ws.append([
            row['table_name'],
            row['column_name'],
            row['original_value'],
            row['display_value'] or '',
            ','.join(row['synonyms'] or []) if row['synonyms'] else '',
            includes_str,
            '是' if is_enabled else '否'
        ])
        # 根据启用状态设置行颜色（只读列1,2保持灰色）
        style_row_by_status(ws, row_idx, is_enabled, len(headers), readonly_cols=[1, 2])
        row_idx += 1
    
    # 设置列宽
    col_widths = [25, 25, 25, 25, 35, 35, 12]
    for i, w in enumerate(col_widths, 1):
        set_column_width(ws, i, w)
    
    # 添加下拉框（使用中文选项）
    data_rows = max(len(enums) + 100, 200)
    add_dropdown(ws, 'G', 2, data_rows, BOOL_OPTIONS_CN, allow_blank=False)
    
    # 添加条件格式：根据"是否启用"列自动变换颜色（只读列1,2保持灰色）
    add_conditional_formatting(ws, 'G', len(headers), data_rows, readonly_cols=[1, 2])
    
    # 设置表格边框
    apply_table_borders(ws, len(headers), row_idx - 1)


async def _create_relationships_sheet(wb: Workbook, db, connection_id: UUID, table_names: Optional[List[str]] = None):
    """创建表关系配置Sheet"""
    ws = wb.create_sheet("表关系配置")
    
    # 表头
    headers = [
        "左表", "左表列", "右表", "右表列",
        "JOIN类型", "关系类型", "关系名称", "描述", "是否启用"
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))
    
    # 构建查询条件
    query = """
        SELECT 
            lt.table_name as left_table, lc.column_name as left_column,
            rt.table_name as right_table, rc.column_name as right_column,
            r.join_type, r.relationship_type, r.relationship_name, r.description, r.is_active
        FROM table_relationships r
        JOIN db_tables lt ON r.left_table_id = lt.table_id
        JOIN db_columns lc ON r.left_column_id = lc.column_id
        JOIN db_tables rt ON r.right_table_id = rt.table_id
        JOIN db_columns rc ON r.right_column_id = rc.column_id
        WHERE r.connection_id = $1
    """
    params = [connection_id]
    
    if table_names:
        query += " AND (lt.table_name = ANY($2::text[]) OR rt.table_name = ANY($2::text[]))"
        params.append(table_names)
    
    query += " ORDER BY lt.table_name, rt.table_name"
    
    relationships = await db.fetch(query, *params)
    
    # 填充数据（使用中文）
    row_idx = 2
    for row in relationships:
        join_type_cn = JOIN_TYPE_EN_TO_CN.get(row['join_type'], row['join_type']) if row['join_type'] else '左连接'
        rel_type_cn = RELATIONSHIP_TYPE_EN_TO_CN.get(row['relationship_type'], row['relationship_type']) if row['relationship_type'] else '一对多'
        is_enabled = row['is_active']
        
        ws.append([
            row['left_table'],
            row['left_column'],
            row['right_table'],
            row['right_column'],
            join_type_cn,
            rel_type_cn,
            row['relationship_name'] or '',
            row['description'] or '',
            '是' if is_enabled else '否'
        ])
        # 根据启用状态设置行颜色
        style_row_by_status(ws, row_idx, is_enabled, len(headers))
        row_idx += 1
    
    # 设置列宽
    col_widths = [25, 20, 25, 20, 12, 15, 30, 35, 12]
    for i, w in enumerate(col_widths, 1):
        set_column_width(ws, i, w)
    
    # 添加下拉框（使用中文选项）
    data_rows = max(len(relationships) + 50, 100)
    add_dropdown(ws, 'E', 2, data_rows, JOIN_TYPE_OPTIONS_CN)
    add_dropdown(ws, 'F', 2, data_rows, RELATIONSHIP_TYPE_OPTIONS_CN)
    add_dropdown(ws, 'I', 2, data_rows, BOOL_OPTIONS_CN, allow_blank=False)
    
    # 添加条件格式：根据"是否启用"列自动变换颜色
    add_conditional_formatting(ws, 'I', len(headers), data_rows)
    
    # 设置表格边框
    apply_table_borders(ws, len(headers), row_idx - 1)


async def _create_rules_sheet(wb: Workbook, db, connection_id: UUID, domain_codes: list):
    """创建全局规则配置Sheet"""
    ws = wb.create_sheet("全局规则配置")
    
    # 表头
    headers = [
        "规则类型", "规则名称", "描述", "规则定义(JSON)", 
        "作用域", "业务域(逗号分隔)", "优先级", "是否启用"
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))
    
    # 获取规则数据
    rules = await db.fetch("""
        SELECT rule_type, rule_name, description, rule_definition,
               scope, domain_ids, priority, is_active
        FROM global_rules
        WHERE connection_id = $1
        ORDER BY rule_type, priority DESC
    """, connection_id)
    
    # 获取域ID到域代码的映射
    domain_id_to_code = {}
    domain_rows = await db.fetch(
        "SELECT domain_id, domain_code FROM business_domains WHERE connection_id = $1",
        connection_id
    )
    for d in domain_rows:
        domain_id_to_code[str(d['domain_id'])] = d['domain_code']
    
    # 填充数据（使用中文）
    row_idx = 2
    for row in rules:
        # 处理规则定义
        rule_def = row['rule_definition']
        if isinstance(rule_def, dict):
            rule_def_str = json.dumps(rule_def, ensure_ascii=False)
        else:
            rule_def_str = str(rule_def) if rule_def else ''
        
        # 处理业务域
        domain_ids = row['domain_ids'] or []
        domain_codes_str = ','.join([
            domain_id_to_code.get(str(did), str(did)) 
            for did in domain_ids
        ])
        
        # 规则类型转中文
        rule_type_cn = RULE_TYPE_EN_TO_CN.get(row['rule_type'], row['rule_type']) if row['rule_type'] else ''
        # 作用域转中文
        scope_cn = '全局' if row['scope'] == 'global' else '业务域'
        is_enabled = row['is_active']
        
        ws.append([
            rule_type_cn,
            row['rule_name'],
            row['description'] or '',
            rule_def_str,
            scope_cn,
            domain_codes_str,
            row['priority'] or 0,
            '是' if is_enabled else '否'
        ])
        # 根据启用状态设置行颜色
        style_row_by_status(ws, row_idx, is_enabled, len(headers))
        row_idx += 1
    
    # 设置列宽
    col_widths = [18, 25, 35, 60, 12, 25, 10, 12]
    for i, w in enumerate(col_widths, 1):
        set_column_width(ws, i, w)
    
    # 添加下拉框（使用中文选项）
    data_rows = max(len(rules) + 50, 100)
    add_dropdown(ws, 'A', 2, data_rows, RULE_TYPE_OPTIONS_CN)
    add_dropdown(ws, 'E', 2, data_rows, ['全局', '业务域'])
    add_dropdown(ws, 'H', 2, data_rows, BOOL_OPTIONS_CN, allow_blank=False)
    
    # 添加条件格式：根据"是否启用"列自动变换颜色
    add_conditional_formatting(ws, 'H', len(headers), data_rows)
    
    # 设置表格边框
    apply_table_borders(ws, len(headers), row_idx - 1)


def _create_instructions_sheet(wb: Workbook):
    """创建说明Sheet"""
    ws = wb.create_sheet("使用说明")
    
    instructions = [
        ["元数据配置模板使用说明"],
        [""],
        ["1. 基本规则"],
        ["   - 灰色背景的列为只读列，由系统自动生成，请勿修改"],
        ["   - 浅绿色背景表示该配置项已启用"],
        ["   - 浅红色背景表示该配置项已禁用"],
        ["   - 带下拉框的列请从下拉选项中选择"],
        [""],
        ["2. 各Sheet说明"],
        ["   - 业务域配置：配置业务域的基本信息和关键词"],
        ["   - 表配置：配置表的显示名称、描述、所属业务域等"],
        ["   - 字段配置：配置字段的语义信息（类型、同义词、单位等）"],
        ["   - 枚举值配置：配置维度字段的枚举值及同义词"],
        ["   - 表关系配置：配置表之间的关联关系"],
        ["   - 全局规则配置：配置派生指标、默认过滤、自定义规则等"],
        [""],
        ["3. 字段类型说明"],
        ["   - 维度：维度字段，用于分组、筛选（如：部门、状态）"],
        ["   - 度量：度量字段，用于聚合计算（如：金额、数量）"],
        ["   - 时间：时间字段（如：创建时间、更新日期）"],
        ["   - 标识：标识字段，如ID、编码等"],
        ["   - 空间：空间字段，地理位置相关"],
        [""],
        ["4. 聚合方式说明（仅度量字段需要配置）"],
        ["   - 求和：对数值求和"],
        ["   - 平均值：计算平均值"],
        ["   - 计数：统计记录数"],
        ["   - 最大值：取最大值"],
        ["   - 最小值：取最小值"],
        ["   - 去重计数：统计不重复的值的数量"],
        [""],
        ["5. 按表导出"],
        ["   - 可以在导出URL中添加 table_name 参数，只导出指定表的配置"],
        ["   - 多个表用逗号分隔，如：table_name=orders,customers"],
        [""],
        ["6. 导入注意事项"],
        ["   - 导入时系统会根据'表名+列名'定位字段"],
        ["   - 新增的枚举值会自动创建"],
        ["   - 已存在的配置会被更新"],
        ["   - 建议先使用'预览'功能查看变更，确认无误后再执行导入"],
    ]
    
    for row in instructions:
        ws.append(row)
    
    # 设置样式
    ws.column_dimensions['A'].width = 80
    ws['A1'].font = Font(bold=True, size=14)


# ============================================================================
# 导入API
# ============================================================================

@router.post("/metadata/import")
async def import_metadata_unified(
    file: UploadFile = File(...),
    connection_ids: Optional[str] = Query(None, description="数据源ID，支持逗号分隔多个；为空时根据Excel中的数据源列自动分发"),
    mode: str = Query("update", description="导入模式: update(更新+新增), merge(仅新增), overwrite(覆盖)"),
    dry_run: bool = Query(False, description="预览模式，不实际执行"),
    current_user: AdminUser = Depends(require_data_admin),
    db = Depends(get_db_pool)
):
    """
    统一的元数据导入端点（支持多数据源）
    
    - 如果Excel来自多数据源导出（有"数据源(只读)"列），自动按数据源名称分发
    - 如果Excel来自单数据源导出，需要指定connection_ids参数
    - connection_ids: 可选，限制只导入到指定的数据源
    """
    try:
        # 验证文件类型
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="请上传Excel文件（.xlsx或.xls格式）"
            )
        
        # 解析数据源ID列表
        connection_id_list = []
        if connection_ids:
            connection_id_list = [UUID(cid.strip()) for cid in connection_ids.split(',') if cid.strip()]
        
        # 读取Excel文件
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        
        # 获取所有数据源的映射：connection_name -> connection_id
        all_connections = await db.fetch(
            "SELECT connection_id, connection_name FROM database_connections WHERE is_active = TRUE"
        )
        conn_name_to_id = {row['connection_name']: row['connection_id'] for row in all_connections}
        
        # 检测Excel格式（是否有"数据源"列）
        is_multi_source = False
        if "表配置" in wb.sheetnames:
            ws = wb["表配置"]
            first_row = [cell.value for cell in ws[1]]
            is_multi_source = first_row and "数据源(只读)" in str(first_row[0])
        elif "业务域配置" in wb.sheetnames:
            ws = wb["业务域配置"]
            first_row = [cell.value for cell in ws[1]]
            is_multi_source = first_row and "数据源(只读)" in str(first_row[0])
        
        # 如果是单数据源格式但没有指定connection_ids，报错
        if not is_multi_source and not connection_id_list:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="该Excel为单数据源格式，请指定connection_ids参数或使用 /metadata/import/{connection_id} 端点"
            )
        
        # 如果是单数据源格式，使用指定的connection_id
        if not is_multi_source:
            if len(connection_id_list) != 1:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="单数据源格式的Excel只能指定一个connection_id"
                )
            # 调用原有的单数据源导入逻辑
            return await _do_import_single(wb, db, connection_id_list[0], mode, dry_run)
        
        # 多数据源格式：按数据源分组导入
        result = {
            "success": True,
            "dry_run": dry_run,
            "by_connection": {},
            "errors": [],
            "warnings": []
        }
        
        # 确定要处理的数据源列表
        if connection_id_list:
            # 只处理指定的数据源
            target_conn_ids = set(connection_id_list)
        else:
            # 从Excel中提取所有涉及的数据源
            target_conn_names = set()
            for sheet_name in wb.sheetnames:
                if sheet_name == "使用说明":
                    continue
                ws = wb[sheet_name]
                first_row = [cell.value for cell in ws[1]]
                if first_row and "数据源" in str(first_row[0]):
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            conn_name = str(row[0]).strip()
                            if conn_name and conn_name != '(全局)':
                                target_conn_names.add(conn_name)
            
            # 转换为connection_id
            target_conn_ids = set()
            for name in target_conn_names:
                if name in conn_name_to_id:
                    target_conn_ids.add(conn_name_to_id[name])
                else:
                    result["warnings"].append({
                        "message": f"数据源 '{name}' 不存在，相关数据将被跳过"
                    })
        
        # 为每个数据源执行导入
        for conn_id in target_conn_ids:
            conn_info = await db.fetchrow(
                "SELECT connection_name FROM database_connections WHERE connection_id = $1",
                conn_id
            )
            if not conn_info:
                result["warnings"].append({
                    "message": f"数据源ID {conn_id} 不存在，已跳过"
                })
                continue
            
            conn_name = conn_info['connection_name']
            
            try:
                # 为该数据源执行导入
                conn_result = await _do_import_multi_source(
                    wb, db, conn_id, conn_name, mode, dry_run, conn_name_to_id
                )
                result["by_connection"][conn_name] = conn_result["summary"]
                result["errors"].extend(conn_result["errors"])
                result["warnings"].extend(conn_result["warnings"])
            except Exception as e:
                result["errors"].append({
                    "connection": conn_name,
                    "message": str(e)
                })
        
        # 如果有错误，标记失败
        if result["errors"]:
            result["success"] = False
        
        logger.info(
            f"导入元数据{'预览' if dry_run else '完成'}（多数据源）",
            mode=mode,
            connections=list(result["by_connection"].keys())
        )
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("导入元数据失败")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"导入失败: {str(e)}"
        )


async def _do_import_single(wb: Workbook, db, connection_id: UUID, mode: str, dry_run: bool) -> dict:
    """执行单数据源导入（原有逻辑）"""
    # 预构建映射表
    mappings = await _build_mappings(db, connection_id)
    
    # 统计结果
    result = {
        "success": True,
        "dry_run": dry_run,
        "summary": {},
        "errors": [],
        "warnings": []
    }
    
    # 按顺序处理各Sheet（依赖顺序很重要）
    if "业务域配置" in wb.sheetnames:
        domains_result = await _import_domains(
            wb["业务域配置"], db, connection_id, mode, dry_run, mappings
        )
        result["summary"]["domains"] = domains_result["stats"]
        result["errors"].extend(domains_result["errors"])
        result["warnings"].extend(domains_result["warnings"])
        if not dry_run:
            mappings = await _build_mappings(db, connection_id)
    
    if "表配置" in wb.sheetnames:
        tables_result = await _import_tables(
            wb["表配置"], db, connection_id, mode, dry_run, mappings
        )
        result["summary"]["tables"] = tables_result["stats"]
        result["errors"].extend(tables_result["errors"])
        result["warnings"].extend(tables_result["warnings"])
    
    if "字段配置" in wb.sheetnames:
        fields_result = await _import_fields(
            wb["字段配置"], db, connection_id, mode, dry_run, mappings
        )
        result["summary"]["fields"] = fields_result["stats"]
        result["errors"].extend(fields_result["errors"])
        result["warnings"].extend(fields_result["warnings"])
    
    if "枚举值配置" in wb.sheetnames:
        enums_result = await _import_enums(
            wb["枚举值配置"], db, connection_id, mode, dry_run, mappings
        )
        result["summary"]["enums"] = enums_result["stats"]
        result["errors"].extend(enums_result["errors"])
        result["warnings"].extend(enums_result["warnings"])
    
    if "表关系配置" in wb.sheetnames:
        rels_result = await _import_relationships(
            wb["表关系配置"], db, connection_id, mode, dry_run, mappings
        )
        result["summary"]["relationships"] = rels_result["stats"]
        result["errors"].extend(rels_result["errors"])
        result["warnings"].extend(rels_result["warnings"])
    
    if "全局规则配置" in wb.sheetnames:
        rules_result = await _import_rules(
            wb["全局规则配置"], db, connection_id, mode, dry_run, mappings
        )
        result["summary"]["rules"] = rules_result["stats"]
        result["errors"].extend(rules_result["errors"])
        result["warnings"].extend(rules_result["warnings"])
    
    if result["errors"]:
        result["success"] = False
    
    return result


async def _do_import_multi_source(wb: Workbook, db, connection_id: UUID, connection_name: str, 
                                   mode: str, dry_run: bool, conn_name_to_id: dict) -> dict:
    """执行多数据源导入（按数据源筛选）"""
    # 预构建映射表
    mappings = await _build_mappings(db, connection_id)
    
    # 统计结果
    result = {
        "success": True,
        "dry_run": dry_run,
        "summary": {},
        "errors": [],
        "warnings": []
    }
    
    # 按顺序处理各Sheet（依赖顺序很重要）
    if "业务域配置" in wb.sheetnames:
        domains_result = await _import_domains_multi(
            wb["业务域配置"], db, connection_id, connection_name, mode, dry_run, mappings
        )
        result["summary"]["domains"] = domains_result["stats"]
        result["errors"].extend(domains_result["errors"])
        result["warnings"].extend(domains_result["warnings"])
        if not dry_run:
            mappings = await _build_mappings(db, connection_id)
    
    if "表配置" in wb.sheetnames:
        tables_result = await _import_tables_multi(
            wb["表配置"], db, connection_id, connection_name, mode, dry_run, mappings
        )
        result["summary"]["tables"] = tables_result["stats"]
        result["errors"].extend(tables_result["errors"])
        result["warnings"].extend(tables_result["warnings"])
    
    if "字段配置" in wb.sheetnames:
        fields_result = await _import_fields_multi(
            wb["字段配置"], db, connection_id, connection_name, mode, dry_run, mappings
        )
        result["summary"]["fields"] = fields_result["stats"]
        result["errors"].extend(fields_result["errors"])
        result["warnings"].extend(fields_result["warnings"])
    
    if "枚举值配置" in wb.sheetnames:
        enums_result = await _import_enums_multi(
            wb["枚举值配置"], db, connection_id, connection_name, mode, dry_run, mappings
        )
        result["summary"]["enums"] = enums_result["stats"]
        result["errors"].extend(enums_result["errors"])
        result["warnings"].extend(enums_result["warnings"])
    
    if "表关系配置" in wb.sheetnames:
        rels_result = await _import_relationships_multi(
            wb["表关系配置"], db, connection_id, connection_name, mode, dry_run, mappings
        )
        result["summary"]["relationships"] = rels_result["stats"]
        result["errors"].extend(rels_result["errors"])
        result["warnings"].extend(rels_result["warnings"])
    
    if "全局规则配置" in wb.sheetnames:
        rules_result = await _import_rules_multi(
            wb["全局规则配置"], db, connection_id, connection_name, mode, dry_run, mappings
        )
        result["summary"]["rules"] = rules_result["stats"]
        result["errors"].extend(rules_result["errors"])
        result["warnings"].extend(rules_result["warnings"])
    
    if result["errors"]:
        result["success"] = False
    
    return result


@router.post("/metadata/import/{connection_id}")
async def import_metadata(
    connection_id: UUID,
    file: UploadFile = File(...),
    mode: str = Query("update", description="导入模式: update(更新+新增), merge(仅新增), overwrite(覆盖)"),
    dry_run: bool = Query(False, description="预览模式，不实际执行"),
    current_user: AdminUser = Depends(require_data_admin),
    db = Depends(get_db_pool)
):
    """
    导入元数据配置（Excel格式）
    
    导入模式：
    - update: 更新已存在的配置，新增不存在的配置（推荐）
    - merge: 仅新增不存在的配置，跳过已存在的
    - overwrite: 覆盖所有配置（谨慎使用）
    
    预览模式：
    - dry_run=true 时只返回预览结果，不实际执行导入
    """
    try:
        # 验证连接存在
        conn_info = await db.fetchrow(
            "SELECT connection_name FROM database_connections WHERE connection_id = $1",
            connection_id
        )
        if not conn_info:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"数据库连接 {connection_id} 不存在"
            )
        
        # 验证文件类型
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="请上传Excel文件（.xlsx或.xls格式）"
            )
        
        # 读取Excel文件
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        
        # 预构建映射表
        mappings = await _build_mappings(db, connection_id)
        
        # 统计结果
        result = {
            "success": True,
            "dry_run": dry_run,
            "summary": {},
            "errors": [],
            "warnings": []
        }
        
        # 按顺序处理各Sheet（依赖顺序很重要）
        if "业务域配置" in wb.sheetnames:
            domains_result = await _import_domains(
                wb["业务域配置"], db, connection_id, mode, dry_run, mappings
            )
            result["summary"]["domains"] = domains_result["stats"]
            result["errors"].extend(domains_result["errors"])
            result["warnings"].extend(domains_result["warnings"])
            # 更新映射（新建的业务域）
            if not dry_run:
                mappings = await _build_mappings(db, connection_id)
        
        if "表配置" in wb.sheetnames:
            tables_result = await _import_tables(
                wb["表配置"], db, connection_id, mode, dry_run, mappings
            )
            result["summary"]["tables"] = tables_result["stats"]
            result["errors"].extend(tables_result["errors"])
            result["warnings"].extend(tables_result["warnings"])
        
        if "字段配置" in wb.sheetnames:
            fields_result = await _import_fields(
                wb["字段配置"], db, connection_id, mode, dry_run, mappings
            )
            result["summary"]["fields"] = fields_result["stats"]
            result["errors"].extend(fields_result["errors"])
            result["warnings"].extend(fields_result["warnings"])
        
        if "枚举值配置" in wb.sheetnames:
            enums_result = await _import_enums(
                wb["枚举值配置"], db, connection_id, mode, dry_run, mappings
            )
            result["summary"]["enums"] = enums_result["stats"]
            result["errors"].extend(enums_result["errors"])
            result["warnings"].extend(enums_result["warnings"])
        
        if "表关系配置" in wb.sheetnames:
            rels_result = await _import_relationships(
                wb["表关系配置"], db, connection_id, mode, dry_run, mappings
            )
            result["summary"]["relationships"] = rels_result["stats"]
            result["errors"].extend(rels_result["errors"])
            result["warnings"].extend(rels_result["warnings"])
        
        if "全局规则配置" in wb.sheetnames:
            rules_result = await _import_rules(
                wb["全局规则配置"], db, connection_id, mode, dry_run, mappings
            )
            result["summary"]["rules"] = rules_result["stats"]
            result["errors"].extend(rules_result["errors"])
            result["warnings"].extend(rules_result["warnings"])
        
        # 如果有错误，标记失败
        if result["errors"]:
            result["success"] = False
        
        logger.info(
            f"导入元数据{'预览' if dry_run else '完成'}",
            connection_id=str(connection_id),
            mode=mode,
            summary=result["summary"]
        )
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("导入元数据失败")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"导入失败: {str(e)}"
        )


async def _build_mappings(db, connection_id: UUID) -> Dict[str, Any]:
    """构建各种映射表，用于导入时定位"""
    mappings = {}
    
    # 业务域映射: domain_code -> domain_id
    # 由于 domain_code 有全局唯一约束，需要获取所有业务域
    # 这样导入时能正确检测已存在的 domain_code 并更新而非报错
    domains = await db.fetch(
        "SELECT domain_id, domain_code, connection_id FROM business_domains"
    )
    mappings['domain_code_to_id'] = {row['domain_code']: row['domain_id'] for row in domains}
    # 记录每个 domain_code 所属的 connection_id，用于判断是否需要更新 connection_id
    mappings['domain_code_to_connection'] = {row['domain_code']: row['connection_id'] for row in domains}
    
    # 表映射: (schema_name, table_name) -> table_id
    tables = await db.fetch(
        "SELECT table_id, schema_name, table_name FROM db_tables WHERE connection_id = $1",
        connection_id
    )
    mappings['table_key_to_id'] = {
        (row['schema_name'] or 'dbo', row['table_name']): row['table_id'] 
        for row in tables
    }
    
    # 列映射: (schema_name, table_name, column_name) -> column_id
    # 同时获取列的详细信息用于自动识别字段类型
    columns = await db.fetch("""
        SELECT c.column_id, t.schema_name, t.table_name, c.column_name,
               c.data_type, c.is_primary_key, c.is_foreign_key
        FROM db_columns c
        JOIN db_tables t ON c.table_id = t.table_id
        WHERE t.connection_id = $1
    """, connection_id)
    mappings['column_key_to_id'] = {
        (row['schema_name'] or 'dbo', row['table_name'], row['column_name']): row['column_id']
        for row in columns
    }
    # 列详细信息: column_id -> {data_type, is_primary_key, is_foreign_key, column_name}
    mappings['column_id_to_info'] = {
        row['column_id']: {
            'column_name': row['column_name'],
            'data_type': row['data_type'] or '',
            'is_primary_key': row['is_primary_key'] or False,
            'is_foreign_key': row['is_foreign_key'] or False
        }
        for row in columns
    }
    
    # 字段映射: column_id -> field_id
    fields = await db.fetch("""
        SELECT f.field_id, f.source_column_id
        FROM fields f
        JOIN db_columns c ON f.source_column_id = c.column_id
        JOIN db_tables t ON c.table_id = t.table_id
        WHERE t.connection_id = $1
    """, connection_id)
    mappings['column_id_to_field_id'] = {row['source_column_id']: row['field_id'] for row in fields}
    
    return mappings


def _parse_list(value: str) -> List[str]:
    """解析逗号分隔的列表"""
    if not value or not str(value).strip():
        return []
    return [v.strip() for v in str(value).split(',') if v.strip()]


def _parse_bool(value) -> bool:
    """解析布尔值"""
    if value is None:
        return True
    if isinstance(value, bool):
        return value
    return str(value).upper() in ('TRUE', '1', 'YES', '是')


async def _import_domains(ws, db, connection_id: UUID, mode: str, dry_run: bool, mappings: dict) -> dict:
    """导入业务域配置"""
    stats = {"new": 0, "update": 0, "skip": 0, "error": 0}
    errors = []
    warnings = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # 跳过空行
            continue
        
        try:
            domain_code = str(row[0]).strip()
            domain_name = str(row[1]).strip() if row[1] else domain_code
            description = str(row[2]).strip() if row[2] else ''
            keywords = _parse_list(row[3])
            typical_queries = _parse_list(row[4])
            icon = str(row[5]).strip() if row[5] else ''
            color = str(row[6]).strip() if row[6] else '#409eff'
            sort_order = int(row[7]) if row[7] else 0
            is_active = _parse_bool(row[8]) if len(row) > 8 else True
            
            existing_id = mappings['domain_code_to_id'].get(domain_code)
            
            if existing_id:
                if mode == 'merge':
                    stats['skip'] += 1
                    continue
                # 更新
                if not dry_run:
                    await db.execute("""
                        UPDATE business_domains 
                        SET domain_name = $1, description = $2, keywords = $3,
                            typical_queries = $4, icon = $5, color = $6, 
                            sort_order = $7, is_active = $8, updated_at = NOW()
                        WHERE domain_id = $9
                    """, domain_name, description, keywords, typical_queries, 
                        icon, color, sort_order, is_active, existing_id)
                stats['update'] += 1
            else:
                # 新增
                if not dry_run:
                    new_id = await db.fetchval("""
                        INSERT INTO business_domains (
                            connection_id, domain_code, domain_name, description,
                            keywords, typical_queries, icon, color, sort_order, is_active
                        ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
                        RETURNING domain_id
                    """, connection_id, domain_code, domain_name, description,
                        keywords, typical_queries, icon, color, sort_order, is_active)
                    mappings['domain_code_to_id'][domain_code] = new_id
                else:
                    # dry_run 模式下也要添加到 mappings，使用临时占位符 UUID
                    # 这样后续 Sheet（如全局规则配置）能正确识别 Excel 中新定义的业务域
                    mappings['domain_code_to_id'][domain_code] = uuid4()
                stats['new'] += 1
                
        except Exception as e:
            stats['error'] += 1
            errors.append({
                "sheet": "业务域配置",
                "row": row_idx,
                "message": str(e)
            })
    
    return {"stats": stats, "errors": errors, "warnings": warnings}


async def _import_tables(ws, db, connection_id: UUID, mode: str, dry_run: bool, mappings: dict) -> dict:
    """导入表配置"""
    stats = {"new": 0, "update": 0, "skip": 0, "error": 0}
    errors = []
    warnings = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        
        try:
            table_name = str(row[0]).strip()
            schema_name = str(row[1]).strip() if row[1] else 'dbo'
            display_name = str(row[2]).strip() if row[2] else ''
            description = str(row[3]).strip() if row[3] else ''
            domain_code = str(row[4]).strip() if row[4] else ''
            tags = _parse_list(row[5])
            data_year = str(row[6]).strip() if row[6] else None
            is_included = _parse_bool(row[7]) if len(row) > 7 else True
            
            table_key = (schema_name, table_name)
            table_id = mappings['table_key_to_id'].get(table_key)
            
            if not table_id:
                warnings.append({
                    "sheet": "表配置",
                    "row": row_idx,
                    "message": f"表 '{schema_name}.{table_name}' 在数据库中不存在，已跳过"
                })
                stats['skip'] += 1
                continue
            
            # 获取业务域ID
            domain_id = mappings['domain_code_to_id'].get(domain_code) if domain_code else None
            
            if mode == 'merge':
                # 检查是否已有配置（display_name不为空）
                existing = await db.fetchval(
                    "SELECT display_name FROM db_tables WHERE table_id = $1",
                    table_id
                )
                if existing:
                    stats['skip'] += 1
                    continue
            
            # 更新表配置
            if not dry_run:
                await db.execute("""
                    UPDATE db_tables 
                    SET display_name = $1, description = $2, domain_id = $3,
                        tags = $4, data_year = $5, is_included = $6, updated_at = NOW()
                    WHERE table_id = $7
                """, display_name or None, description or None, domain_id,
                    tags, data_year, is_included, table_id)
            stats['update'] += 1
            
        except Exception as e:
            stats['error'] += 1
            errors.append({
                "sheet": "表配置",
                "row": row_idx,
                "message": str(e)
            })
    
    return {"stats": stats, "errors": errors, "warnings": warnings}


async def _import_fields(ws, db, connection_id: UUID, mode: str, dry_run: bool, mappings: dict) -> dict:
    """导入字段配置（支持中文选项）"""
    stats = {"new": 0, "update": 0, "skip": 0, "error": 0}
    errors = []
    warnings = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        
        try:
            table_name = str(row[0]).strip()
            schema_name = str(row[1]).strip() if row[1] else 'dbo'
            column_name = str(row[2]).strip()
            # row[3] 是数据类型（只读）
            display_name = str(row[4]).strip() if row[4] else ''
            field_type_raw = str(row[5]).strip() if row[5] else ''
            # row[6] 是同义词（维度类型已移除）
            synonyms = _parse_list(row[6])
            default_aggregation_raw = str(row[7]).strip() if row[7] else ''
            unit = str(row[8]).strip() if row[8] else ''
            is_active = _parse_bool(row[9]) if len(row) > 9 else True
            show_in_detail = _parse_bool(row[10]) if len(row) > 10 else False
            description = str(row[11]).strip() if len(row) > 11 and row[11] else ''
            
            # 中文转英文
            field_type = FIELD_TYPE_CN_TO_EN.get(field_type_raw, field_type_raw) if field_type_raw else ''
            default_aggregation = AGGREGATION_CN_TO_EN.get(default_aggregation_raw, default_aggregation_raw) if default_aggregation_raw else ''
            
            # 定位列
            column_key = (schema_name, table_name, column_name)
            column_id = mappings['column_key_to_id'].get(column_key)
            
            if not column_id:
                warnings.append({
                    "sheet": "字段配置",
                    "row": row_idx,
                    "message": f"列 '{schema_name}.{table_name}.{column_name}' 在数据库中不存在，已跳过"
                })
                stats['skip'] += 1
                continue
            
            # 检查是否已有field记录
            field_id = mappings['column_id_to_field_id'].get(column_id)
            
            # 验证字段类型（支持中英文）
            valid_field_types = list(FIELD_TYPE_CN_TO_EN.keys()) + list(FIELD_TYPE_CN_TO_EN.values())
            if field_type and field_type not in valid_field_types and field_type_raw not in valid_field_types:
                errors.append({
                    "sheet": "字段配置",
                    "row": row_idx,
                    "column": "字段类型",
                    "message": f"无效的字段类型: {field_type_raw}，有效值: {FIELD_TYPE_OPTIONS_CN}"
                })
                stats['error'] += 1
                continue

            # 如果字段类型为空，自动识别
            auto_detected = False
            if not field_type:
                col_info = mappings.get('column_id_to_info', {}).get(column_id)
                if col_info:
                    analysis = FieldAnalyzer.analyze_field(
                        column_name=col_info['column_name'],
                        data_type=col_info['data_type'],
                        is_primary_key=col_info['is_primary_key'],
                        is_foreign_key=col_info['is_foreign_key']
                    )
                    field_type = analysis.field_type
                    # 如果用户没指定显示名称，使用分析器生成的
                    if not display_name:
                        display_name = analysis.display_name
                    # 如果是度量类型且用户没指定聚合方式和单位，使用分析器推荐的
                    if field_type == 'measure':
                        if not default_aggregation and analysis.default_aggregation:
                            default_aggregation = analysis.default_aggregation.lower()
                        if not unit and analysis.unit:
                            unit = analysis.unit
                    auto_detected = True
                else:
                    field_type = 'dimension'

            if field_id:
                if mode == 'merge':
                    stats['skip'] += 1
                    continue
                # 更新
                if not dry_run:
                    await db.execute("""
                        UPDATE fields
                        SET display_name = $1, field_type = $2,
                            synonyms = $3, default_aggregation = $4, unit = $5,
                            is_active = $6, show_in_detail = $7, description = $8,
                            auto_detected = $10, updated_at = NOW()
                        WHERE field_id = $9
                    """, display_name or column_name, field_type,
                        synonyms, default_aggregation or None, unit or None,
                        is_active, show_in_detail, description or None, field_id, auto_detected)
                stats['update'] += 1
            else:
                # 新增
                if not dry_run:
                    new_field_id = await db.fetchval("""
                        INSERT INTO fields (
                            connection_id, source_type, source_column_id,
                            display_name, field_type, synonyms,
                            default_aggregation, unit, is_active, show_in_detail,
                            description, auto_detected
                        )
                        SELECT t.connection_id, 'column', $1, $2, $3, $4, $5, $6, $7, $8, $9, $10
                        FROM db_columns c
                        JOIN db_tables t ON c.table_id = t.table_id
                        WHERE c.column_id = $1
                        RETURNING field_id
                    """, column_id, display_name or column_name, field_type,
                        synonyms, default_aggregation or None,
                        unit or None, is_active, show_in_detail, description or None, auto_detected)
                    mappings['column_id_to_field_id'][column_id] = new_field_id
                stats['new'] += 1
                
        except Exception as e:
            stats['error'] += 1
            errors.append({
                "sheet": "字段配置",
                "row": row_idx,
                "message": str(e)
            })
    
    return {"stats": stats, "errors": errors, "warnings": warnings}


async def _import_enums(ws, db, connection_id: UUID, mode: str, dry_run: bool, mappings: dict) -> dict:
    """导入枚举值配置"""
    stats = {"new": 0, "update": 0, "skip": 0, "error": 0}
    errors = []
    warnings = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0] or not row[2]:  # 表名和原始值必须有
            continue
        
        try:
            table_name = str(row[0]).strip()
            column_name = str(row[1]).strip()
            original_value = str(row[2]).strip()
            display_value = str(row[3]).strip() if row[3] else ''
            synonyms = _parse_list(row[4])
            includes_values = _parse_list(row[5])
            is_active = _parse_bool(row[6]) if len(row) > 6 else True
            
            # 定位字段（需要通过表名+列名找到field_id）
            # 先找column_id
            column_key = ('dbo', table_name, column_name)
            column_id = mappings['column_key_to_id'].get(column_key)
            
            if not column_id:
                # 尝试不带schema
                for key, cid in mappings['column_key_to_id'].items():
                    if key[1] == table_name and key[2] == column_name:
                        column_id = cid
                        break
            
            if not column_id:
                warnings.append({
                    "sheet": "枚举值配置",
                    "row": row_idx,
                    "message": f"列 '{table_name}.{column_name}' 在数据库中不存在，已跳过"
                })
                stats['skip'] += 1
                continue
            
            # 获取field_id
            field_id = mappings['column_id_to_field_id'].get(column_id)

            if not field_id:
                # 字段不存在，需要先创建字段（自动识别类型）
                if not dry_run:
                    col_info = mappings.get('column_id_to_info', {}).get(column_id)
                    if col_info:
                        analysis = FieldAnalyzer.analyze_field(
                            column_name=col_info['column_name'],
                            data_type=col_info['data_type'],
                            is_primary_key=col_info['is_primary_key'],
                            is_foreign_key=col_info['is_foreign_key']
                        )
                        field_type = analysis.field_type
                        display_name_auto = analysis.display_name
                    else:
                        field_type = 'dimension'
                        display_name_auto = column_name

                    field_id = await db.fetchval("""
                        INSERT INTO fields (
                            connection_id, source_type, source_column_id,
                            display_name, field_type, auto_detected, is_active, show_in_detail
                        )
                        SELECT t.connection_id, 'column', $1, $2, $3, TRUE, TRUE, FALSE
                        FROM db_columns c
                        JOIN db_tables t ON c.table_id = t.table_id
                        WHERE c.column_id = $1
                        RETURNING field_id
                    """, column_id, display_name_auto, field_type)
                    mappings['column_id_to_field_id'][column_id] = field_id
            
            # 检查枚举值是否已存在
            existing_enum = None
            if not dry_run:
                existing_enum = await db.fetchval("""
                    SELECT enum_value_id FROM field_enum_values 
                    WHERE field_id = $1 AND original_value = $2
                """, field_id, original_value)
            
            if existing_enum:
                if mode == 'merge':
                    stats['skip'] += 1
                    continue
                # 更新
                if not dry_run:
                    await db.execute("""
                        UPDATE field_enum_values 
                        SET display_value = $1, synonyms = $2, includes_values = $3,
                            is_active = $4, updated_at = NOW()
                        WHERE enum_value_id = $5
                    """, display_value or None, synonyms, includes_values or None, is_active, existing_enum)
                stats['update'] += 1
            else:
                # 新增
                if not dry_run:
                    await db.execute("""
                        INSERT INTO field_enum_values (
                            field_id, original_value, display_value, synonyms,
                            includes_values, is_active
                        ) VALUES ($1, $2, $3, $4, $5, $6)
                    """, field_id, original_value, display_value or None, synonyms,
                        includes_values or None, is_active)
                stats['new'] += 1
                
        except Exception as e:
            stats['error'] += 1
            errors.append({
                "sheet": "枚举值配置",
                "row": row_idx,
                "message": str(e)
            })
    
    return {"stats": stats, "errors": errors, "warnings": warnings}


async def _import_relationships(ws, db, connection_id: UUID, mode: str, dry_run: bool, mappings: dict) -> dict:
    """导入表关系配置（支持中文选项）"""
    stats = {"new": 0, "update": 0, "skip": 0, "error": 0}
    errors = []
    warnings = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0] or not row[2]:  # 左表和右表必须有
            continue
        
        try:
            left_table = str(row[0]).strip()
            left_column = str(row[1]).strip()
            right_table = str(row[2]).strip()
            right_column = str(row[3]).strip()
            join_type_raw = str(row[4]).strip() if row[4] else '左连接'
            relationship_type_raw = str(row[5]).strip() if row[5] else '一对多'
            
            # 中文转英文
            join_type = JOIN_TYPE_CN_TO_EN.get(join_type_raw, join_type_raw.upper()) if join_type_raw else 'LEFT'
            relationship_type = RELATIONSHIP_TYPE_CN_TO_EN.get(relationship_type_raw, relationship_type_raw) if relationship_type_raw else 'one_to_many'
            relationship_name = str(row[6]).strip() if row[6] else ''
            description = str(row[7]).strip() if row[7] else ''
            is_active = _parse_bool(row[8]) if len(row) > 8 else True
            
            # 定位表和列
            left_table_id = None
            left_column_id = None
            right_table_id = None
            right_column_id = None
            
            for key, tid in mappings['table_key_to_id'].items():
                if key[1] == left_table:
                    left_table_id = tid
                if key[1] == right_table:
                    right_table_id = tid
            
            for key, cid in mappings['column_key_to_id'].items():
                if key[1] == left_table and key[2] == left_column:
                    left_column_id = cid
                if key[1] == right_table and key[2] == right_column:
                    right_column_id = cid
            
            if not all([left_table_id, left_column_id, right_table_id, right_column_id]):
                warnings.append({
                    "sheet": "表关系配置",
                    "row": row_idx,
                    "message": f"关系 '{left_table}.{left_column} -> {right_table}.{right_column}' 中存在无效的表或列，已跳过"
                })
                stats['skip'] += 1
                continue
            
            # 检查关系是否已存在
            existing_rel = None
            if not dry_run:
                existing_rel = await db.fetchval("""
                    SELECT relationship_id FROM table_relationships 
                    WHERE connection_id = $1 
                      AND left_table_id = $2 AND left_column_id = $3
                      AND right_table_id = $4 AND right_column_id = $5
                """, connection_id, left_table_id, left_column_id, right_table_id, right_column_id)
            
            if existing_rel:
                if mode == 'merge':
                    stats['skip'] += 1
                    continue
                # 更新
                if not dry_run:
                    await db.execute("""
                        UPDATE table_relationships 
                        SET join_type = $1, relationship_type = $2, relationship_name = $3,
                            description = $4, is_active = $5, updated_at = NOW()
                        WHERE relationship_id = $6
                    """, join_type, relationship_type, relationship_name or None,
                        description or None, is_active, existing_rel)
                stats['update'] += 1
            else:
                # 新增
                if not dry_run:
                    await db.execute("""
                        INSERT INTO table_relationships (
                            connection_id, left_table_id, right_table_id,
                            left_column_id, right_column_id, join_type,
                            relationship_type, relationship_name, description,
                            detection_method, is_confirmed, is_active
                        ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, 'manual', TRUE, $10)
                    """, connection_id, left_table_id, right_table_id,
                        left_column_id, right_column_id, join_type,
                        relationship_type, relationship_name or None, description or None, is_active)
                stats['new'] += 1
                
        except Exception as e:
            stats['error'] += 1
            errors.append({
                "sheet": "表关系配置",
                "row": row_idx,
                "message": str(e)
            })
    
    return {"stats": stats, "errors": errors, "warnings": warnings}


async def _import_rules(ws, db, connection_id: UUID, mode: str, dry_run: bool, mappings: dict) -> dict:
    """导入全局规则配置（支持中文选项）"""
    stats = {"new": 0, "update": 0, "skip": 0, "error": 0}
    errors = []
    warnings = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0] or not row[1]:  # 规则类型和规则名称必须有
            continue
        
        try:
            rule_type_raw = str(row[0]).strip()
            rule_name = str(row[1]).strip()
            description = str(row[2]).strip() if row[2] else ''
            rule_definition_str = str(row[3]).strip() if row[3] else '{}'
            scope_raw = str(row[4]).strip() if row[4] else '全局'
            domain_codes_str = str(row[5]).strip() if row[5] else ''
            priority = int(row[6]) if row[6] else 0
            is_active = _parse_bool(row[7]) if len(row) > 7 else True
            
            # 中文转英文
            rule_type = RULE_TYPE_CN_TO_EN.get(rule_type_raw, rule_type_raw) if rule_type_raw else ''
            scope = 'global' if scope_raw in ('全局', 'global') else 'domain'
            
            # 验证规则类型（支持中英文）
            valid_rule_types = list(RULE_TYPE_CN_TO_EN.keys()) + list(RULE_TYPE_CN_TO_EN.values())
            if rule_type not in valid_rule_types and rule_type_raw not in valid_rule_types:
                errors.append({
                    "sheet": "全局规则配置",
                    "row": row_idx,
                    "column": "规则类型",
                    "message": f"无效的规则类型: {rule_type_raw}，有效值: {RULE_TYPE_OPTIONS_CN}"
                })
                stats['error'] += 1
                continue
            
            # 解析规则定义
            try:
                rule_definition = json.loads(rule_definition_str)
            except json.JSONDecodeError:
                errors.append({
                    "sheet": "全局规则配置",
                    "row": row_idx,
                    "column": "规则定义",
                    "message": f"无效的JSON格式: {rule_definition_str[:50]}..."
                })
                stats['error'] += 1
                continue
            
            # 解析业务域
            domain_ids = []
            if domain_codes_str:
                for code in _parse_list(domain_codes_str):
                    did = mappings['domain_code_to_id'].get(code)
                    if did:
                        domain_ids.append(did)
                    else:
                        warnings.append({
                            "sheet": "全局规则配置",
                            "row": row_idx,
                            "message": f"业务域代码 '{code}' 不存在，已忽略"
                        })
            
            # 检查规则是否已存在（全局规则按 rule_type + rule_name 唯一）
            existing_rule = None
            if not dry_run:
                # 全局规则不绑定 connection_id，按类型+名称查找
                existing_rule = await db.fetchval("""
                    SELECT rule_id FROM global_rules 
                    WHERE rule_type = $1 AND rule_name = $2
                """, rule_type, rule_name)
            
            if existing_rule:
                if mode == 'merge':
                    stats['skip'] += 1
                    continue
                # 更新
                if not dry_run:
                    await db.execute("""
                        UPDATE global_rules 
                        SET description = $1, rule_definition = $2::jsonb, scope = $3,
                            domain_ids = $4, priority = $5, is_active = $6, updated_at = NOW()
                        WHERE rule_id = $7
                    """, description or None, json.dumps(rule_definition, ensure_ascii=False),
                        scope, domain_ids or None, priority, is_active, existing_rule)
                stats['update'] += 1
            else:
                # 新增（全局规则，connection_id 设为 NULL）
                if not dry_run:
                    await db.execute("""
                        INSERT INTO global_rules (
                            connection_id, rule_type, rule_name, description,
                            rule_definition, scope, domain_ids, priority, is_active
                        ) VALUES (NULL, $1, $2, $3, $4::jsonb, $5, $6, $7, $8)
                    """, rule_type, rule_name, description or None,
                        json.dumps(rule_definition, ensure_ascii=False), scope,
                        domain_ids or None, priority, is_active)
                stats['new'] += 1
                
        except Exception as e:
            stats['error'] += 1
            errors.append({
                "sheet": "全局规则配置",
                "row": row_idx,
                "message": str(e)
            })
    
    return {"stats": stats, "errors": errors, "warnings": warnings}


# ============================================================================
# 多数据源导入辅助函数（按数据源列筛选）
# ============================================================================

async def _import_domains_multi(ws, db, connection_id: UUID, connection_name: str, mode: str, dry_run: bool, mappings: dict) -> dict:
    """导入业务域配置（多数据源格式，按数据源筛选）"""
    stats = {"new": 0, "update": 0, "skip": 0, "error": 0}
    errors = []
    warnings = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0] or not row[1]:  # 跳过空行
            continue
        
        # 检查是否属于当前数据源
        row_conn_name = str(row[0]).strip()
        if row_conn_name != connection_name and row_conn_name != '(全局)':
            continue
        
        try:
            # 多数据源格式：列偏移1
            domain_code = str(row[1]).strip()
            domain_name = str(row[2]).strip() if row[2] else domain_code
            description = str(row[3]).strip() if row[3] else ''
            keywords = _parse_list(row[4])
            typical_queries = _parse_list(row[5])
            icon = str(row[6]).strip() if row[6] else ''
            color = str(row[7]).strip() if row[7] else '#409eff'
            sort_order = int(row[8]) if row[8] else 0
            is_active = _parse_bool(row[9]) if len(row) > 9 else True
            
            existing_id = mappings['domain_code_to_id'].get(domain_code)
            
            if existing_id:
                if mode == 'merge':
                    stats['skip'] += 1
                    continue
                if not dry_run:
                    await db.execute("""
                        UPDATE business_domains 
                        SET domain_name = $1, description = $2, keywords = $3,
                            typical_queries = $4, icon = $5, color = $6, 
                            sort_order = $7, is_active = $8, updated_at = NOW()
                        WHERE domain_id = $9
                    """, domain_name, description, keywords, typical_queries, 
                        icon, color, sort_order, is_active, existing_id)
                stats['update'] += 1
            else:
                if not dry_run:
                    new_id = await db.fetchval("""
                        INSERT INTO business_domains (
                            connection_id, domain_code, domain_name, description,
                            keywords, typical_queries, icon, color, sort_order, is_active
                        ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
                        RETURNING domain_id
                    """, connection_id, domain_code, domain_name, description,
                        keywords, typical_queries, icon, color, sort_order, is_active)
                    mappings['domain_code_to_id'][domain_code] = new_id
                else:
                    # dry_run 模式下也要添加到 mappings，使用临时占位符 UUID
                    # 这样后续 Sheet（如全局规则配置）能正确识别 Excel 中新定义的业务域
                    mappings['domain_code_to_id'][domain_code] = uuid4()
                stats['new'] += 1
                
        except Exception as e:
            stats['error'] += 1
            errors.append({"sheet": "业务域配置", "row": row_idx, "message": str(e)})
    
    return {"stats": stats, "errors": errors, "warnings": warnings}


async def _import_tables_multi(ws, db, connection_id: UUID, connection_name: str, mode: str, dry_run: bool, mappings: dict) -> dict:
    """导入表配置（多数据源格式，按数据源筛选）"""
    stats = {"new": 0, "update": 0, "skip": 0, "error": 0}
    errors = []
    warnings = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0] or not row[1]:
            continue
        
        row_conn_name = str(row[0]).strip()
        if row_conn_name != connection_name:
            continue
        
        try:
            # 多数据源格式：列偏移1
            table_name = str(row[1]).strip()
            schema_name = str(row[2]).strip() if row[2] else 'dbo'
            display_name = str(row[3]).strip() if row[3] else ''
            description = str(row[4]).strip() if row[4] else ''
            domain_code = str(row[5]).strip() if row[5] else ''
            tags = _parse_list(row[6])
            data_year = str(row[7]).strip() if row[7] else None
            is_included = _parse_bool(row[8]) if len(row) > 8 else True
            
            table_key = (schema_name, table_name)
            table_id = mappings['table_key_to_id'].get(table_key)
            
            if not table_id:
                warnings.append({
                    "sheet": "表配置", "row": row_idx,
                    "message": f"表 '{schema_name}.{table_name}' 在数据库中不存在，已跳过"
                })
                stats['skip'] += 1
                continue
            
            domain_id = mappings['domain_code_to_id'].get(domain_code) if domain_code else None
            
            if mode == 'merge':
                existing = await db.fetchval("SELECT display_name FROM db_tables WHERE table_id = $1", table_id)
                if existing:
                    stats['skip'] += 1
                    continue
            
            if not dry_run:
                await db.execute("""
                    UPDATE db_tables 
                    SET display_name = $1, description = $2, domain_id = $3,
                        tags = $4, data_year = $5, is_included = $6, updated_at = NOW()
                    WHERE table_id = $7
                """, display_name or None, description or None, domain_id, tags, data_year, is_included, table_id)
            stats['update'] += 1
            
        except Exception as e:
            stats['error'] += 1
            errors.append({"sheet": "表配置", "row": row_idx, "message": str(e)})
    
    return {"stats": stats, "errors": errors, "warnings": warnings}


async def _import_fields_multi(ws, db, connection_id: UUID, connection_name: str, mode: str, dry_run: bool, mappings: dict) -> dict:
    """导入字段配置（多数据源格式，按数据源筛选）"""
    stats = {"new": 0, "update": 0, "skip": 0, "error": 0}
    errors = []
    warnings = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0] or not row[1]:
            continue
        
        row_conn_name = str(row[0]).strip()
        if row_conn_name != connection_name:
            continue
        
        try:
            # 多数据源格式：数据源、表名、Schema、列名、数据类型、显示名称、字段类型、同义词、默认聚合、单位、是否启用、显示在明细、描述
            table_name = str(row[1]).strip()
            schema_name = str(row[2]).strip() if row[2] else 'dbo'
            column_name = str(row[3]).strip()
            # row[4] 是数据类型（只读）
            display_name = str(row[5]).strip() if row[5] else ''
            field_type_raw = str(row[6]).strip() if row[6] else ''
            synonyms = _parse_list(row[7])
            default_aggregation_raw = str(row[8]).strip() if row[8] else ''
            unit = str(row[9]).strip() if row[9] else ''
            is_active = _parse_bool(row[10]) if len(row) > 10 else True
            show_in_detail = _parse_bool(row[11]) if len(row) > 11 else False
            description = str(row[12]).strip() if len(row) > 12 and row[12] else ''
            
            # 中文转英文
            field_type = FIELD_TYPE_CN_TO_EN.get(field_type_raw, field_type_raw) if field_type_raw else ''
            default_aggregation = AGGREGATION_CN_TO_EN.get(default_aggregation_raw, default_aggregation_raw) if default_aggregation_raw else ''
            
            column_key = (schema_name, table_name, column_name)
            column_id = mappings['column_key_to_id'].get(column_key)
            
            if not column_id:
                warnings.append({
                    "sheet": "字段配置", "row": row_idx,
                    "message": f"列 '{schema_name}.{table_name}.{column_name}' 不存在，已跳过"
                })
                stats['skip'] += 1
                continue
            
            field_id = mappings['column_id_to_field_id'].get(column_id)

            # 如果字段类型为空，自动识别
            auto_detected = False
            if not field_type:
                col_info = mappings.get('column_id_to_info', {}).get(column_id)
                if col_info:
                    analysis = FieldAnalyzer.analyze_field(
                        column_name=col_info['column_name'],
                        data_type=col_info['data_type'],
                        is_primary_key=col_info['is_primary_key'],
                        is_foreign_key=col_info['is_foreign_key']
                    )
                    field_type = analysis.field_type
                    if not display_name:
                        display_name = analysis.display_name
                    if field_type == 'measure':
                        if not default_aggregation and analysis.default_aggregation:
                            default_aggregation = analysis.default_aggregation.lower()
                        if not unit and analysis.unit:
                            unit = analysis.unit
                    auto_detected = True
                else:
                    field_type = 'dimension'

            if field_id:
                if mode == 'merge':
                    stats['skip'] += 1
                    continue
                if not dry_run:
                    await db.execute("""
                        UPDATE fields
                        SET display_name = $1, field_type = $2, synonyms = $3,
                            default_aggregation = $4, unit = $5, is_active = $6,
                            show_in_detail = $7, description = $8, auto_detected = $10, updated_at = NOW()
                        WHERE field_id = $9
                    """, display_name or column_name, field_type, synonyms,
                        default_aggregation or None, unit or None, is_active, show_in_detail,
                        description or None, field_id, auto_detected)
                stats['update'] += 1
            else:
                if not dry_run:
                    new_field_id = await db.fetchval("""
                        INSERT INTO fields (
                            connection_id, source_type, source_column_id, display_name,
                            field_type, synonyms, default_aggregation, unit,
                            is_active, show_in_detail, description, auto_detected
                        )
                        SELECT t.connection_id, 'column', $1, $2, $3, $4, $5, $6, $7, $8, $9, $10
                        FROM db_columns c JOIN db_tables t ON c.table_id = t.table_id
                        WHERE c.column_id = $1
                        RETURNING field_id
                    """, column_id, display_name or column_name, field_type,
                        synonyms, default_aggregation or None, unit or None,
                        is_active, show_in_detail, description or None, auto_detected)
                    mappings['column_id_to_field_id'][column_id] = new_field_id
                stats['new'] += 1
                
        except Exception as e:
            stats['error'] += 1
            errors.append({"sheet": "字段配置", "row": row_idx, "message": str(e)})
    
    return {"stats": stats, "errors": errors, "warnings": warnings}


async def _import_enums_multi(ws, db, connection_id: UUID, connection_name: str, mode: str, dry_run: bool, mappings: dict) -> dict:
    """导入枚举值配置（多数据源格式，按数据源筛选）"""
    stats = {"new": 0, "update": 0, "skip": 0, "error": 0}
    errors = []
    warnings = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0] or not row[3]:  # 数据源和原始值必须有
            continue
        
        row_conn_name = str(row[0]).strip()
        if row_conn_name != connection_name:
            continue
        
        try:
            # 多数据源格式：数据源、表名、列名、原始值、显示值、同义词、包含值、是否启用
            table_name = str(row[1]).strip()
            column_name = str(row[2]).strip()
            original_value = str(row[3]).strip()
            display_value = str(row[4]).strip() if row[4] else ''
            synonyms = _parse_list(row[5])
            includes_values = _parse_list(row[6])
            is_active = _parse_bool(row[7]) if len(row) > 7 else True
            
            # 定位字段
            column_key = ('dbo', table_name, column_name)
            column_id = mappings['column_key_to_id'].get(column_key)
            
            if not column_id:
                for key, cid in mappings['column_key_to_id'].items():
                    if key[1] == table_name and key[2] == column_name:
                        column_id = cid
                        break
            
            if not column_id:
                warnings.append({
                    "sheet": "枚举值配置", "row": row_idx,
                    "message": f"列 '{table_name}.{column_name}' 不存在，已跳过"
                })
                stats['skip'] += 1
                continue
            
            field_id = mappings['column_id_to_field_id'].get(column_id)

            if not field_id:
                # 字段不存在，需要先创建字段（自动识别类型）
                if not dry_run:
                    col_info = mappings.get('column_id_to_info', {}).get(column_id)
                    if col_info:
                        analysis = FieldAnalyzer.analyze_field(
                            column_name=col_info['column_name'],
                            data_type=col_info['data_type'],
                            is_primary_key=col_info['is_primary_key'],
                            is_foreign_key=col_info['is_foreign_key']
                        )
                        field_type = analysis.field_type
                        display_name_auto = analysis.display_name
                    else:
                        field_type = 'dimension'
                        display_name_auto = column_name

                    field_id = await db.fetchval("""
                        INSERT INTO fields (connection_id, source_type, source_column_id, display_name, field_type, auto_detected, is_active, show_in_detail)
                        SELECT t.connection_id, 'column', $1, $2, $3, TRUE, TRUE, FALSE
                        FROM db_columns c JOIN db_tables t ON c.table_id = t.table_id WHERE c.column_id = $1
                        RETURNING field_id
                    """, column_id, display_name_auto, field_type)
                    mappings['column_id_to_field_id'][column_id] = field_id
            
            existing_enum = None
            if not dry_run and field_id:
                existing_enum = await db.fetchval(
                    "SELECT enum_value_id FROM field_enum_values WHERE field_id = $1 AND original_value = $2",
                    field_id, original_value
                )
            
            if existing_enum:
                if mode == 'merge':
                    stats['skip'] += 1
                    continue
                if not dry_run:
                    await db.execute("""
                        UPDATE field_enum_values 
                        SET display_value = $1, synonyms = $2, includes_values = $3, is_active = $4, updated_at = NOW()
                        WHERE enum_value_id = $5
                    """, display_value or None, synonyms, includes_values or None, is_active, existing_enum)
                stats['update'] += 1
            else:
                if not dry_run and field_id:
                    await db.execute("""
                        INSERT INTO field_enum_values (field_id, original_value, display_value, synonyms, includes_values, is_active)
                        VALUES ($1, $2, $3, $4, $5, $6)
                    """, field_id, original_value, display_value or None, synonyms, includes_values or None, is_active)
                stats['new'] += 1
                
        except Exception as e:
            stats['error'] += 1
            errors.append({"sheet": "枚举值配置", "row": row_idx, "message": str(e)})
    
    return {"stats": stats, "errors": errors, "warnings": warnings}


async def _import_relationships_multi(ws, db, connection_id: UUID, connection_name: str, mode: str, dry_run: bool, mappings: dict) -> dict:
    """导入表关系配置（多数据源格式，按数据源筛选）"""
    stats = {"new": 0, "update": 0, "skip": 0, "error": 0}
    errors = []
    warnings = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0] or not row[1]:
            continue
        
        row_conn_name = str(row[0]).strip()
        if row_conn_name != connection_name:
            continue
        
        try:
            # 多数据源格式：数据源、左表、左表列、右表、右表列、JOIN类型、关系类型、关系名称、描述、是否启用
            left_table = str(row[1]).strip()
            left_column = str(row[2]).strip()
            right_table = str(row[3]).strip()
            right_column = str(row[4]).strip()
            join_type_raw = str(row[5]).strip() if row[5] else '左连接'
            relationship_type_raw = str(row[6]).strip() if row[6] else '一对多'
            relationship_name = str(row[7]).strip() if row[7] else ''
            description = str(row[8]).strip() if row[8] else ''
            is_active = _parse_bool(row[9]) if len(row) > 9 else True
            
            join_type = JOIN_TYPE_CN_TO_EN.get(join_type_raw, join_type_raw.upper()) if join_type_raw else 'LEFT'
            relationship_type = RELATIONSHIP_TYPE_CN_TO_EN.get(relationship_type_raw, relationship_type_raw) if relationship_type_raw else 'one_to_many'
            
            # 定位表和列
            left_table_id = left_column_id = right_table_id = right_column_id = None
            
            for key, tid in mappings['table_key_to_id'].items():
                if key[1] == left_table:
                    left_table_id = tid
                if key[1] == right_table:
                    right_table_id = tid
            
            for key, cid in mappings['column_key_to_id'].items():
                if key[1] == left_table and key[2] == left_column:
                    left_column_id = cid
                if key[1] == right_table and key[2] == right_column:
                    right_column_id = cid
            
            if not all([left_table_id, left_column_id, right_table_id, right_column_id]):
                warnings.append({
                    "sheet": "表关系配置", "row": row_idx,
                    "message": f"关系 '{left_table}.{left_column} -> {right_table}.{right_column}' 中存在无效的表或列，已跳过"
                })
                stats['skip'] += 1
                continue
            
            existing_rel = None
            if not dry_run:
                existing_rel = await db.fetchval("""
                    SELECT relationship_id FROM table_relationships 
                    WHERE connection_id = $1 AND left_table_id = $2 AND left_column_id = $3
                      AND right_table_id = $4 AND right_column_id = $5
                """, connection_id, left_table_id, left_column_id, right_table_id, right_column_id)
            
            if existing_rel:
                if mode == 'merge':
                    stats['skip'] += 1
                    continue
                if not dry_run:
                    await db.execute("""
                        UPDATE table_relationships 
                        SET join_type = $1, relationship_type = $2, relationship_name = $3,
                            description = $4, is_active = $5, updated_at = NOW()
                        WHERE relationship_id = $6
                    """, join_type, relationship_type, relationship_name or None, description or None, is_active, existing_rel)
                stats['update'] += 1
            else:
                if not dry_run:
                    await db.execute("""
                        INSERT INTO table_relationships (
                            connection_id, left_table_id, right_table_id, left_column_id, right_column_id,
                            join_type, relationship_type, relationship_name, description, detection_method, is_confirmed, is_active
                        ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, 'manual', TRUE, $10)
                    """, connection_id, left_table_id, right_table_id, left_column_id, right_column_id,
                        join_type, relationship_type, relationship_name or None, description or None, is_active)
                stats['new'] += 1
                
        except Exception as e:
            stats['error'] += 1
            errors.append({"sheet": "表关系配置", "row": row_idx, "message": str(e)})
    
    return {"stats": stats, "errors": errors, "warnings": warnings}


async def _import_rules_multi(ws, db, connection_id: UUID, connection_name: str, mode: str, dry_run: bool, mappings: dict) -> dict:
    """导入全局规则配置（多数据源格式）
    
    注意：全局规则配置Sheet没有数据源列，所有规则都会导入到指定的connection_id
    """
    # 直接复用原有的单数据源导入逻辑
    return await _import_rules(ws, db, connection_id, mode, dry_run, mappings)


# ============================================================================
# 导出所有/多数据源的辅助函数
# ============================================================================

async def _create_domains_sheet_all(wb: Workbook, db, domain_rows: list, connection_ids: Optional[List[UUID]] = None):
    """创建业务域配置Sheet（所有/多个数据源）"""
    ws = wb.create_sheet("业务域配置")
    
    # 表头（添加数据源列）
    headers = ["数据源(只读)", "域代码", "域名称", "描述", "关键词(逗号分隔)", "典型查询(逗号分隔)", "图标", "颜色", "排序", "是否启用"]
    ws.append(headers)
    style_header_row(ws, len(headers))
    
    # 获取业务域（包含连接信息）
    if connection_ids:
        rows = await db.fetch("""
            SELECT bd.*, dc.connection_name 
            FROM business_domains bd
            LEFT JOIN database_connections dc ON bd.connection_id = dc.connection_id
            WHERE bd.connection_id = ANY($1) OR bd.connection_id IS NULL
            ORDER BY dc.connection_name NULLS FIRST, bd.sort_order
        """, connection_ids)
    else:
        rows = await db.fetch("""
            SELECT bd.*, dc.connection_name 
            FROM business_domains bd
            LEFT JOIN database_connections dc ON bd.connection_id = dc.connection_id
            ORDER BY dc.connection_name NULLS FIRST, bd.sort_order
        """)
    
    row_idx = 2
    for row in rows:
        is_active = row['is_active']
        ws.append([
            row['connection_name'] or '(全局)',
            row['domain_code'],
            row['domain_name'],
            row['description'] or '',
            ','.join(row['keywords'] or []),
            ','.join(row['typical_queries'] or []),
            row['icon'] or '',
            row['color'] or '#409eff',
            row['sort_order'] or 0,
            '是' if is_active else '否'
        ])
        # 设置只读列和状态颜色
        style_row_by_status(ws, row_idx, is_active, len(headers), readonly_cols=[1])
        row_idx += 1
    
    # 设置列宽
    col_widths = [20, 15, 20, 30, 30, 40, 15, 12, 8, 12]
    for i, w in enumerate(col_widths, 1):
        set_column_width(ws, i, w)
    
    # 添加下拉框
    data_rows = len(rows) + 100
    add_dropdown(ws, 'J', 2, data_rows, BOOL_OPTIONS_CN, allow_blank=False)
    
    # 添加条件格式
    add_conditional_formatting(ws, 'J', len(headers), data_rows, readonly_cols=[1])
    
    # 设置表格边框
    apply_table_borders(ws, len(headers), row_idx - 1)


async def _create_tables_sheet_all(wb: Workbook, db, domain_codes: list, domain_map: dict, table_names: Optional[List[str]] = None, connection_ids: Optional[List[UUID]] = None):
    """创建表配置Sheet（所有/多个数据源）"""
    ws = wb.create_sheet("表配置")
    
    # 表头（添加数据源列）
    headers = ["数据源(只读)", "表名(只读)", "Schema(只读)", "显示名称", "描述", "所属业务域", "标签(逗号分隔)", "数据年份", "是否启用"]
    ws.append(headers)
    style_header_row(ws, len(headers))
    
    # 构建查询
    query = """
        SELECT t.*, dc.connection_name, bd.domain_code
        FROM db_tables t
        JOIN database_connections dc ON t.connection_id = dc.connection_id
        LEFT JOIN business_domains bd ON t.domain_id = bd.domain_id
        WHERE 1=1
    """
    params = []
    param_idx = 1
    
    if connection_ids:
        query += f" AND t.connection_id = ANY(${param_idx})"
        params.append(connection_ids)
        param_idx += 1
    
    if table_names:
        query += f" AND t.table_name = ANY(${param_idx}::text[])"
        params.append(table_names)
    
    query += " ORDER BY dc.connection_name, t.schema_name, t.table_name"
    
    rows = await db.fetch(query, *params)
    
    row_idx = 2
    for row in rows:
        is_enabled = row['is_included']
        ws.append([
            row['connection_name'],
            row['table_name'],
            row['schema_name'] or 'dbo',
            row['display_name'] or '',
            row['description'] or '',
            row['domain_code'] or '',
            ','.join(row['tags'] or []),
            row['data_year'] or '',
            '是' if is_enabled else '否'
        ])
        # 设置只读列和状态颜色
        style_row_by_status(ws, row_idx, is_enabled, len(headers), readonly_cols=[1, 2, 3])
        row_idx += 1
    
    # 设置列宽
    col_widths = [20, 25, 15, 20, 35, 15, 25, 12, 12]
    for i, w in enumerate(col_widths, 1):
        set_column_width(ws, i, w)
    
    # 添加下拉框
    data_rows = len(rows) + 100
    if domain_codes:
        add_dropdown(ws, 'F', 2, data_rows, domain_codes)
    add_dropdown(ws, 'I', 2, data_rows, BOOL_OPTIONS_CN, allow_blank=False)
    
    # 添加条件格式
    add_conditional_formatting(ws, 'I', len(headers), data_rows, readonly_cols=[1, 2, 3])
    
    # 设置表格边框
    apply_table_borders(ws, len(headers), row_idx - 1)


async def _create_fields_sheet_all(wb: Workbook, db, table_names: Optional[List[str]] = None, connection_ids: Optional[List[UUID]] = None):
    """创建字段配置Sheet（所有/多个数据源）"""
    ws = wb.create_sheet("字段配置")
    
    # 表头（添加数据源列）
    headers = [
        "数据源(只读)", "表名(只读)", "Schema(只读)", "列名(只读)", "数据类型(只读)",
        "显示名称", "字段类型", "同义词(逗号分隔)",
        "默认聚合", "单位", "是否启用", "显示在明细", "描述"
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))
    
    # 构建查询 - 使用正确的字段关联
    query = """
        SELECT 
            dc.connection_name, t.table_name, t.schema_name, c.column_name, c.data_type,
            f.display_name, f.field_type, f.synonyms,
            f.default_aggregation, f.unit, f.is_active, f.show_in_detail, f.description
        FROM db_tables t
        JOIN database_connections dc ON t.connection_id = dc.connection_id
        JOIN db_columns c ON t.table_id = c.table_id
        LEFT JOIN fields f ON c.column_id = f.source_column_id
        WHERE t.is_included = TRUE
    """
    params = []
    param_idx = 1
    
    if connection_ids:
        query += f" AND t.connection_id = ANY(${param_idx})"
        params.append(connection_ids)
        param_idx += 1
    
    if table_names:
        query += f" AND t.table_name = ANY(${param_idx}::text[])"
        params.append(table_names)
    
    query += " ORDER BY dc.connection_name, t.schema_name, t.table_name, c.ordinal_position NULLS LAST"
    
    rows = await db.fetch(query, *params)
    
    row_idx = 2
    for row in rows:
        # 字段类型转中文
        field_type_cn = FIELD_TYPE_EN_TO_CN.get(row['field_type'], '') if row['field_type'] else ''
        # 聚合方式转中文
        agg_cn = AGGREGATION_EN_TO_CN.get(row['default_aggregation'], '') if row['default_aggregation'] else ''
        # 启用状态
        is_enabled = row['is_active'] is None or row['is_active']
        
        ws.append([
            row['connection_name'],
            row['table_name'],
            row['schema_name'] or 'dbo',
            row['column_name'],
            row['data_type'],
            row['display_name'] or '',
            field_type_cn,
            ','.join(row['synonyms'] or []) if row['synonyms'] else '',
            agg_cn,
            row['unit'] or '',
            '是' if is_enabled else '否',
            '是' if row['show_in_detail'] is None or row['show_in_detail'] else '否',
            row['description'] or ''
        ])
        # 设置只读列和状态颜色
        style_row_by_status(ws, row_idx, is_enabled, len(headers), readonly_cols=[1, 2, 3, 4, 5])
        row_idx += 1
    
    # 设置列宽
    col_widths = [20, 25, 12, 25, 15, 18, 12, 30, 14, 10, 12, 12, 35]
    for i, w in enumerate(col_widths, 1):
        set_column_width(ws, i, w)
    
    # 添加下拉框
    data_rows = len(rows) + 100
    add_dropdown(ws, 'G', 2, data_rows, FIELD_TYPE_OPTIONS_CN)  # 字段类型
    add_dropdown(ws, 'I', 2, data_rows, AGGREGATION_OPTIONS_CN)  # 默认聚合
    add_dropdown(ws, 'K', 2, data_rows, BOOL_OPTIONS_CN, allow_blank=False)  # 是否启用
    add_dropdown(ws, 'L', 2, data_rows, BOOL_OPTIONS_CN, allow_blank=False)  # 显示在明细
    
    # 添加条件格式
    add_conditional_formatting(ws, 'K', len(headers), data_rows, readonly_cols=[1, 2, 3, 4, 5])
    
    # 设置表格边框
    apply_table_borders(ws, len(headers), row_idx - 1)


async def _create_enums_sheet_all(wb: Workbook, db, table_names: Optional[List[str]] = None, connection_ids: Optional[List[UUID]] = None):
    """创建枚举值配置Sheet（所有/多个数据源）"""
    ws = wb.create_sheet("枚举值配置")
    
    # 表头（添加数据源列）
    headers = [
        "数据源(只读)", "表名(只读)", "列名(只读)", "原始值", "显示值", 
        "同义词(逗号分隔)", "包含值(逗号分隔)", "是否启用"
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))
    
    # 构建查询 - 使用正确的字段关联
    query = """
        SELECT 
            dc.connection_name, t.table_name, c.column_name,
            e.original_value, e.display_value, e.synonyms, e.includes_values, e.is_active
        FROM field_enum_values e
        JOIN fields f ON e.field_id = f.field_id
        JOIN db_columns c ON f.source_column_id = c.column_id
        JOIN db_tables t ON c.table_id = t.table_id
        JOIN database_connections dc ON t.connection_id = dc.connection_id
        WHERE 1=1
    """
    params = []
    param_idx = 1
    
    if connection_ids:
        query += f" AND t.connection_id = ANY(${param_idx})"
        params.append(connection_ids)
        param_idx += 1
    
    if table_names:
        query += f" AND t.table_name = ANY(${param_idx}::text[])"
        params.append(table_names)
    
    query += " ORDER BY dc.connection_name, t.table_name, c.column_name, e.frequency DESC"
    
    rows = await db.fetch(query, *params)
    
    row_idx = 2
    for row in rows:
        includes_values = row['includes_values']
        if includes_values and isinstance(includes_values, list):
            includes_str = ','.join(includes_values)
        else:
            includes_str = ''
        
        is_enabled = row['is_active']
        ws.append([
            row['connection_name'],
            row['table_name'],
            row['column_name'],
            row['original_value'],
            row['display_value'] or '',
            ','.join(row['synonyms'] or []) if row['synonyms'] else '',
            includes_str,
            '是' if is_enabled else '否'
        ])
        # 设置只读列和状态颜色
        style_row_by_status(ws, row_idx, is_enabled, len(headers), readonly_cols=[1, 2, 3])
        row_idx += 1
    
    # 设置列宽
    col_widths = [20, 25, 25, 25, 25, 35, 35, 12]
    for i, w in enumerate(col_widths, 1):
        set_column_width(ws, i, w)
    
    # 添加下拉框
    data_rows = max(len(rows) + 100, 200)
    add_dropdown(ws, 'H', 2, data_rows, BOOL_OPTIONS_CN, allow_blank=False)
    
    # 添加条件格式
    add_conditional_formatting(ws, 'H', len(headers), data_rows, readonly_cols=[1, 2, 3])
    
    # 设置表格边框
    apply_table_borders(ws, len(headers), row_idx - 1)


async def _create_relationships_sheet_all(wb: Workbook, db, table_names: Optional[List[str]] = None, connection_ids: Optional[List[UUID]] = None):
    """创建表关系配置Sheet（所有/多个数据源）"""
    ws = wb.create_sheet("表关系配置")
    
    # 表头（添加数据源列）
    headers = [
        "数据源(只读)", "左表", "左表列", "右表", "右表列",
        "JOIN类型", "关系类型", "关系名称", "描述", "是否启用"
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))
    
    # 构建查询
    query = """
        SELECT dc.connection_name,
               lt.table_name as left_table, lc.column_name as left_column,
               rt.table_name as right_table, rc.column_name as right_column,
               r.join_type, r.relationship_type, r.relationship_name, r.description, r.is_active
        FROM table_relationships r
        JOIN database_connections dc ON r.connection_id = dc.connection_id
        JOIN db_tables lt ON r.left_table_id = lt.table_id
        JOIN db_columns lc ON r.left_column_id = lc.column_id
        JOIN db_tables rt ON r.right_table_id = rt.table_id
        JOIN db_columns rc ON r.right_column_id = rc.column_id
        WHERE 1=1
    """
    params = []
    param_idx = 1
    
    if connection_ids:
        query += f" AND r.connection_id = ANY(${param_idx})"
        params.append(connection_ids)
        param_idx += 1
    
    if table_names:
        query += f" AND (lt.table_name = ANY(${param_idx}::text[]) OR rt.table_name = ANY(${param_idx}::text[]))"
        params.append(table_names)
    
    query += " ORDER BY dc.connection_name, lt.table_name, rt.table_name"
    
    rows = await db.fetch(query, *params)
    
    row_idx = 2
    for row in rows:
        join_type_cn = JOIN_TYPE_EN_TO_CN.get(row['join_type'], row['join_type']) if row['join_type'] else '左连接'
        rel_type_cn = RELATIONSHIP_TYPE_EN_TO_CN.get(row['relationship_type'], row['relationship_type']) if row['relationship_type'] else '一对多'
        is_enabled = row['is_active']
        
        ws.append([
            row['connection_name'],
            row['left_table'],
            row['left_column'],
            row['right_table'],
            row['right_column'],
            join_type_cn,
            rel_type_cn,
            row['relationship_name'] or '',
            row['description'] or '',
            '是' if is_enabled else '否'
        ])
        # 设置只读列和状态颜色
        style_row_by_status(ws, row_idx, is_enabled, len(headers), readonly_cols=[1])
        row_idx += 1
    
    # 设置列宽
    col_widths = [20, 25, 20, 25, 20, 12, 15, 30, 35, 12]
    for i, w in enumerate(col_widths, 1):
        set_column_width(ws, i, w)
    
    # 添加下拉框
    data_rows = max(len(rows) + 50, 100)
    add_dropdown(ws, 'F', 2, data_rows, JOIN_TYPE_OPTIONS_CN)
    add_dropdown(ws, 'G', 2, data_rows, RELATIONSHIP_TYPE_OPTIONS_CN)
    add_dropdown(ws, 'J', 2, data_rows, BOOL_OPTIONS_CN, allow_blank=False)
    
    # 添加条件格式
    add_conditional_formatting(ws, 'J', len(headers), data_rows, readonly_cols=[1])
    
    # 设置表格边框
    apply_table_borders(ws, len(headers), row_idx - 1)


async def _create_rules_sheet_all(wb: Workbook, db, domain_codes: list, connection_ids: Optional[List[UUID]] = None):
    """创建全局规则配置Sheet（所有/多个数据源的规则）"""
    ws = wb.create_sheet("全局规则配置")
    
    # 表头
    headers = [
        "规则类型", "规则名称", "描述", "规则定义(JSON)", 
        "作用域", "业务域(逗号分隔)", "优先级", "是否启用"
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))
    
    # 构建查询
    query = """
        SELECT rule_type, rule_name, description, rule_definition,
               scope, domain_ids, priority, is_active
        FROM global_rules
        WHERE 1=1
    """
    params = []
    
    if connection_ids:
        query += " AND connection_id = ANY($1)"
        params.append(connection_ids)
    
    query += " ORDER BY rule_type, priority DESC"
    
    rules = await db.fetch(query, *params)
    
    # 获取域ID到域代码的映射
    domain_id_to_code = {}
    if connection_ids:
        domain_rows = await db.fetch(
            "SELECT domain_id, domain_code FROM business_domains WHERE connection_id = ANY($1) OR connection_id IS NULL",
            connection_ids
        )
    else:
        domain_rows = await db.fetch("SELECT domain_id, domain_code FROM business_domains")
    for d in domain_rows:
        domain_id_to_code[str(d['domain_id'])] = d['domain_code']
    
    # 填充数据
    row_idx = 2
    for row in rules:
        # 处理规则定义
        rule_def = row['rule_definition']
        if isinstance(rule_def, dict):
            rule_def_str = json.dumps(rule_def, ensure_ascii=False)
        else:
            rule_def_str = str(rule_def) if rule_def else ''
        
        # 处理业务域
        domain_ids = row['domain_ids'] or []
        domain_codes_str = ','.join([
            domain_id_to_code.get(str(did), str(did)) 
            for did in domain_ids
        ])
        
        # 规则类型转中文
        rule_type_cn = RULE_TYPE_EN_TO_CN.get(row['rule_type'], row['rule_type']) if row['rule_type'] else ''
        # 作用域转中文
        scope_cn = '全局' if row['scope'] == 'global' else '业务域'
        is_enabled = row['is_active']
        
        ws.append([
            rule_type_cn,
            row['rule_name'],
            row['description'] or '',
            rule_def_str,
            scope_cn,
            domain_codes_str,
            row['priority'] or 0,
            '是' if is_enabled else '否'
        ])
        # 设置状态颜色
        style_row_by_status(ws, row_idx, is_enabled, len(headers))
        row_idx += 1
    
    # 设置列宽
    col_widths = [18, 25, 35, 60, 12, 25, 10, 12]
    for i, w in enumerate(col_widths, 1):
        set_column_width(ws, i, w)
    
    # 添加下拉框
    data_rows = max(len(rules) + 50, 100)
    add_dropdown(ws, 'A', 2, data_rows, RULE_TYPE_OPTIONS_CN)
    add_dropdown(ws, 'E', 2, data_rows, ['全局', '业务域'])
    add_dropdown(ws, 'H', 2, data_rows, BOOL_OPTIONS_CN, allow_blank=False)
    
    # 添加条件格式
    add_conditional_formatting(ws, 'H', len(headers), data_rows)
    
    # 设置表格边框
    apply_table_borders(ws, len(headers), row_idx - 1)
